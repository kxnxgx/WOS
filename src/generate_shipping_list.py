"""
generate_shipping_list.py
--------------------------
暫定移動明細.xlsx + 商品マスタ.csv から
  1. transfer_upload.csv（システム取込用縦持ちCSV / 大分類列付き）
  2. WOS_Report_店舗出荷.xlsx（店舗別出荷指示Excel）
     - シート1: 店舗別出荷サマリー（9×9マトリクス + 大分類別サマリー）
     - シート2: 統合移動明細（全明細・大分類列）
     - シート3〜11: 出荷_<店舗名>（全9店舗別指示書）
を生成するスクリプト。

実行方法:
    python src/generate_shipping_list.py
    python src/generate_shipping_list.py --input 暫定移動明細.xlsx --master 商品マスタ.csv
"""

import os
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# store list
ALL_STORES = [
    'TOKYO', 'ルクア大阪', '名古屋', 'TOKYO NODE',
    '京王新宿', '大丸心斎橋', '玉川高島屋', 'HUTTE', 'NARITA'
]


def get_store_area(store_name):
    s = str(store_name).strip()
    if 'NODE' in s:                    return 'KANTO'
    elif 'TOKYO' in s:                 return 'KANTO'
    elif '新宿' in s:                   return 'KANTO'
    elif '高島屋' in s or '玉川' in s:  return 'KANTO'
    elif 'NARITA' in s:                return 'KANTO'
    elif 'ルクア' in s:                 return 'KANSAI'
    elif '心斎橋' in s or '大丸' in s:  return 'KANSAI'
    elif '名古屋' in s:                 return 'TOKAI'
    elif 'HUT' in s:                   return 'HOKKAIDO'
    return 'OTHER'


def get_transfer_cost(shipper, receiver):
    a1 = get_store_area(shipper)
    a2 = get_store_area(receiver)
    if a1 == a2 and a1 != 'OTHER':
        return 1
    pair = {a1, a2}
    if pair in ({'KANTO', 'TOKAI'}, {'KANSAI', 'TOKAI'}):
        return 2
    elif pair == {'KANTO', 'KANSAI'}:
        return 3
    elif 'HOKKAIDO' in pair:
        return 4
    return 5


def parse_target_store(comment_text, candidate_stores):
    if not comment_text:
        return None
    c = comment_text.strip().lower()
    for store in candidate_stores:
        if store.lower() in c:
            return store
    keywords = [
        ('node', 'TOKYO NODE'), ('tokyo', 'TOKYO'),
        ('ルクア', 'ルクア大阪'), ('心斎橋', '大丸心斎橋'), ('大丸', '大丸心斎橋'),
        ('名古屋', '名古屋'), ('新宿', '京王新宿'), ('京王', '京王新宿'),
        ('高島屋', '玉川高島屋'), ('玉川', '玉川高島屋'),
        ('hut', 'HUTTE'), ('narita', 'NARITA'),
    ]
    for kw, target_name in keywords:
        if kw in c:
            for store in candidate_stores:
                if target_name in store or store in target_name:
                    return store
    return None


def setup_styles():
    thin = Side(border_style='thin', color='D9D9D9')
    dbl  = Side(border_style='double', color='000000')
    return {
        'font_main':      Font(name='Meiryo UI', size=9),
        'font_bold':      Font(name='Meiryo UI', size=9, bold=True),
        'font_title':     Font(name='Meiryo UI', size=13, bold=True, color='1F4E79'),
        'font_subtitle':  Font(name='Meiryo UI', size=9.5, bold=True, color='595959'),
        'font_header':    Font(name='Meiryo UI', size=9, bold=True, color='FFFFFF'),
        'font_header_dk': Font(name='Meiryo UI', size=9, bold=True, color='000000'),
        'font_red_bold':  Font(name='Meiryo UI', size=11, bold=True, color='C00000'),
        'font_qty_blue':  Font(name='Meiryo UI', size=11, bold=True, color='002060'),
        'fill_navy':      PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
        'fill_blue':      PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
        'fill_dark':      PatternFill(start_color='333F48', end_color='333F48', fill_type='solid'),
        'fill_yellow':    PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'fill_qty_green': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'fill_total':     PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'fill_apparel':   PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid'),
        'fill_hardwear':  PatternFill(start_color='E8FFE8', end_color='E8FFE8', fill_type='solid'),
        'fill_bulk_org':  PatternFill(start_color='FFEAA7', end_color='FFEAA7', fill_type='solid'),
        'border_thin':    Border(left=thin, right=thin, top=thin, bottom=thin),
        'border_total':   Border(left=thin, right=thin, top=thin, bottom=dbl),
        'align_c':        Alignment(horizontal='center', vertical='center'),
        'align_l':        Alignment(horizontal='left',   vertical='center'),
        'align_r':        Alignment(horizontal='right',  vertical='center'),
    }


def wos_str(wos_days_val):
    try:
        days = float(wos_days_val)
        if days < 0 or days > 9990:
            return '-'
        return f"{days / 7:.1f}週"
    except (TypeError, ValueError):
        return '-'


def _cell(ws, row, col, value, st, font='font_main', fill=None, align='align_l', border='border_thin'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = st[font]
    cell.alignment = st[align]
    cell.border    = st[border]
    if fill:
        cell.fill = st[fill]
    return cell


def load_and_match(input_path, master_path):
    print("[1/4] 商品マスタ.csv を読み込んでいます...")
    df_m = pd.read_csv(master_path, encoding='cp932', usecols=[0, 23], dtype=str)
    df_m.columns = ['sku', 'major_category']
    master_cat = (df_m.dropna(subset=['major_category'])
                  .drop_duplicates('sku')
                  .set_index('sku')['major_category']
                  .to_dict())

    print("[2/4] 暫定移動明細.xlsx を読み込んでいます...")
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb['order'] if 'order' in wb.sheetnames else wb.active

    row1 = [c.value for c in ws[1]]
    row2 = [c.value for c in ws[2]]

    sku_col = name_col = color_col = st_col = bulk_col = None
    order_cols = []
    stock_cols = []
    wos_cols   = []

    for idx in range(1, len(row1) + 1):
        cat  = str(row1[idx-1] or '').strip()
        item = str(row2[idx-1] or '').strip()
        if item == '商品コード':        sku_col   = idx
        elif item == '商品名':          name_col  = idx
        elif item == 'カラー':          color_col = idx
        elif '消化率' in item:          st_col    = idx
        elif item == 'BULK在庫':        bulk_col  = idx
        elif cat.upper() == 'ORDER':    order_cols.append((idx, item))
        elif cat.upper() == '在庫':     stock_cols.append((idx, item))
        elif cat.upper() == '在庫日数': wos_cols.append((idx, item))

    if not sku_col:
        raise ValueError("商品コード列が見つかりません")
    if not order_cols:
        raise ValueError("ORDER列が見つかりません")

    candidate_stores = [s for _, s in order_cols]
    stock_map = {s: idx for idx, s in stock_cols}
    wos_map   = {s: idx for idx, s in wos_cols}
    print(f"    対象店舗: {len(candidate_stores)} ({', '.join(candidate_stores)})")

    store_stock = {s: {} for s in candidate_stores}
    store_wos   = {s: {} for s in candidate_stores}
    item_info   = {}

    print("[3/4] データ行を解析・マッチング中...")
    records = []
    unbalanced = 0

    for r in range(3, ws.max_row + 1):
        sku_val = ws.cell(row=r, column=sku_col).value
        if not sku_val or str(sku_val).strip() == '':
            continue

        sku        = str(sku_val).strip()
        item_name  = str(ws.cell(row=r, column=name_col).value  or '').strip() if name_col  else ''
        color_name = str(ws.cell(row=r, column=color_col).value or '').strip() if color_col else ''
        sell_thru  = ws.cell(row=r, column=st_col).value   if st_col   else None
        bulk_raw   = ws.cell(row=r, column=bulk_col).value if bulk_col else 0
        bulk_qty   = int(bulk_raw) if isinstance(bulk_raw, (int, float)) else 0
        major_cat  = master_cat.get(sku, '未分類')

        item_info[sku] = {
            'name': item_name, 'color': color_name,
            'bulk_stock': bulk_qty, 'sell_through': sell_thru,
            'major_cat': major_cat,
        }

        for sn, si in stock_map.items():
            v = ws.cell(row=r, column=si).value
            store_stock[sn][sku] = int(v) if isinstance(v, (int, float)) else 0
        for sn, wi in wos_map.items():
            store_wos[sn][sku] = ws.cell(row=r, column=wi).value

        shippers  = []
        receivers = []
        for ci, sname in order_cols:
            cell = ws.cell(row=r, column=ci)
            v    = cell.value
            cmt  = cell.comment.text if cell.comment else None
            spec = parse_target_store(cmt, candidate_stores)
            if v is not None and isinstance(v, (int, float)):
                vi = int(v)
                if vi < 0:
                    shippers.append({'store': sname, 'qty': abs(vi), 'target': spec})
                elif vi > 0:
                    receivers.append({'store': sname, 'qty': vi, 'source': spec})

        if not shippers and not receivers:
            continue

        if sum(s['qty'] for s in shippers) != sum(r2['qty'] for r2 in receivers):
            unbalanced += 1

        # コメント指定マッチング
        for s in shippers:
            if s['target'] and s['qty'] > 0:
                for r2 in receivers:
                    if r2['store'] == s['target'] and r2['qty'] > 0:
                        alloc = min(s['qty'], r2['qty'])
                        if alloc > 0:
                            records.append(_rec(s['store'], r2['store'], sku, item_info[sku], alloc, '📌コメント指定'))
                            s['qty'] -= alloc; r2['qty'] -= alloc
        for r2 in receivers:
            if r2.get('source') and r2['qty'] > 0:
                for s in shippers:
                    if s['store'] == r2['source'] and s['qty'] > 0:
                        alloc = min(s['qty'], r2['qty'])
                        if alloc > 0:
                            records.append(_rec(s['store'], r2['store'], sku, item_info[sku], alloc, '📌コメント指定'))
                            s['qty'] -= alloc; r2['qty'] -= alloc

        # コスト最小化マッチング
        pairs = []
        for s in shippers:
            if s['qty'] <= 0: continue
            for r2 in receivers:
                if r2['qty'] <= 0: continue
                pairs.append((get_transfer_cost(s['store'], r2['store']), s, r2))
        pairs.sort(key=lambda x: x[0])

        for cost, s, r2 in pairs:
            if s['qty'] <= 0 or r2['qty'] <= 0: continue
            alloc = min(s['qty'], r2['qty'])
            if alloc > 0:
                label = '🚚同一エリア' if cost == 1 else ('🚛隣接エリア' if cost == 2 else '✈️長距離移動')
                records.append(_rec(s['store'], r2['store'], sku, item_info[sku], alloc, label))
                s['qty'] -= alloc; r2['qty'] -= alloc

    if unbalanced:
        print(f"    [警告] {unbalanced} SKU で出入庫不一致")
    total_pts = sum(r['数量'] for r in records)
    print(f"    マッチング完了: {len(records)} 件 / {total_pts} 点")
    return records, store_stock, store_wos, item_info


def _rec(from_s, to_s, sku, info, qty, match_type):
    return {
        '大分類': info['major_cat'], '移動元店舗': from_s, '移動先店舗': to_s,
        '商品コード': sku, '商品名': info['name'], 'カラー': info['color'],
        '数量': qty, '消化率(%)': info['sell_through'],
        'BULK在庫': info['bulk_stock'], 'マッチング種別': match_type,
    }


def export_csv(records, output_path):
    print("[4a] transfer_upload.csv を出力...")
    df = pd.DataFrame(records)
    cols = ['大分類', '移動元店舗', '移動先店舗', '商品コード', '商品名', 'カラー', '数量']
    df_out = df[[c for c in cols if c in df.columns]].sort_values(
        ['大分類', '移動元店舗', '移動先店舗', '商品コード'])
    df_out.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"    → {output_path} ({len(df_out)} 件 / {int(df['数量'].sum())} 点)")


def export_excel(records, store_stock, store_wos, item_info, output_path):
    print("[4b] WOS_Report_店舗出荷.xlsx を生成...")
    df_all = pd.DataFrame(records)

    df_all['出荷前WOS']  = df_all.apply(
        lambda row: wos_str(store_wos.get(row['移動元店舗'], {}).get(row['商品コード'])), axis=1)
    df_all['受入前WOS']  = df_all.apply(
        lambda row: wos_str(store_wos.get(row['移動先店舗'], {}).get(row['商品コード'])), axis=1)
    df_all['出荷元在庫'] = df_all.apply(
        lambda row: store_stock.get(row['移動元店舗'], {}).get(row['商品コード'], 0), axis=1)
    df_all['受入先在庫'] = df_all.apply(
        lambda row: store_stock.get(row['移動先店舗'], {}).get(row['商品コード'], 0), axis=1)

    st = setup_styles()
    wb = openpyxl.Workbook()
    default_ws = wb.active

    print("    生成中: 店舗別出荷サマリー")
    ws1 = wb.create_sheet(title='店舗別出荷サマリー', index=0)
    build_summary_sheet(ws1, df_all, st)

    print("    生成中: 統合移動明細")
    ws2 = wb.create_sheet(title='統合移動明細', index=1)
    build_detail_sheet(ws2, df_all, st)

    for store_name in ALL_STORES:
        print(f"    生成中: 出荷_{store_name}")
        ws_s = wb.create_sheet(title=f'出荷_{store_name}')
        build_store_sheet(ws_s, df_all, store_name, st)

    wb.remove(default_ws)
    wb.save(output_path)
    print(f"    → {output_path} (シート数: {len(wb.sheetnames)})")


CAT_FILL = {'Apparel': 'fill_apparel', 'Hardwear': 'fill_hardwear'}


def build_summary_sheet(ws, df_all, st):
    ws.views.sheetView[0].showGridLines = True
    ws.column_dimensions['A'].width = 3

    total_items = len(df_all)
    total_qty   = int(df_all['数量'].sum())
    categories  = sorted(df_all['大分類'].unique().tolist())

    ws['B2'] = f"店舗間移動 出荷・受入サマリー（合計 {total_items} 件 / {total_qty} 点）"
    ws['B2'].font = st['font_title']
    ws['B3'] = "※ 各店舗が出荷する点数および受入店舗の内訳一覧です。"
    ws['B3'].font = st['font_subtitle']

    # ── 9×9 マトリクス ──
    matrix = df_all.groupby(['移動元店舗', '移動先店舗'])['数量'].sum().unstack(fill_value=0)
    matrix = matrix.reindex(index=ALL_STORES, columns=ALL_STORES, fill_value=0)

    sr = 5
    sc = 2
    _cell(ws, sr, sc, '出荷元 \\ 受入先', st, font='font_header', fill='fill_dark', align='align_c', border='border_thin')
    for ci, recv in enumerate(ALL_STORES, sc+1):
        _cell(ws, sr, ci, recv, st, font='font_header', fill='fill_dark', align='align_c', border='border_thin')
    tcol = sc + len(ALL_STORES) + 1
    icol = tcol + 1
    _cell(ws, sr, tcol, '出荷合計', st, font='font_header', fill='fill_navy', align='align_c', border='border_thin')
    _cell(ws, sr, icol, '出荷件数', st, font='font_header', fill='fill_navy', align='align_c', border='border_thin')

    for ri, ship in enumerate(ALL_STORES, sr+1):
        _cell(ws, ri, sc, ship, st, font='font_bold', fill='fill_total', align='align_c', border='border_thin')
        row_tot = 0
        for ci, recv in enumerate(ALL_STORES, sc+1):
            qty = int(matrix.loc[ship, recv]) if ship in matrix.index and recv in matrix.columns else 0
            row_tot += qty
            _cell(ws, ri, ci, qty if qty > 0 else '-', st,
                  font='font_bold' if qty > 0 else 'font_main',
                  fill='fill_qty_green' if qty > 0 else None,
                  align='align_c', border='border_thin')
        _cell(ws, ri, tcol, row_tot, st,
              font='font_bold', fill='fill_yellow' if row_tot > 0 else None,
              align='align_c', border='border_thin')
        ic = len(df_all[df_all['移動元店舗'] == ship])
        _cell(ws, ri, icol, f'{ic}件' if ic > 0 else '-', st,
              font='font_bold' if ic > 0 else 'font_main', align='align_c', border='border_thin')

    bot = sr + len(ALL_STORES) + 1
    _cell(ws, bot, sc, '受入合計', st, font='font_header', fill='fill_navy', align='align_c', border='border_total')
    grand = 0
    for ci, recv in enumerate(ALL_STORES, sc+1):
        ct = int(matrix[recv].sum()) if recv in matrix.columns else 0
        grand += ct
        _cell(ws, bot, ci, ct if ct > 0 else '-', st,
              font='font_bold', fill='fill_yellow' if ct > 0 else None,
              align='align_c', border='border_total')
    _cell(ws, bot, tcol, grand, st, font='font_red_bold', fill='fill_yellow', align='align_c', border='border_total')
    _cell(ws, bot, icol, f'{total_items}件', st, font='font_red_bold', fill='fill_yellow', align='align_c', border='border_total')

    ws.column_dimensions[get_column_letter(sc)].width = 22
    for ci in range(sc+1, icol+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── 大分類別サマリー ──
    cat_title_r = bot + 3
    ws.cell(row=cat_title_r-1, column=sc, value="大分類別 出荷・受入内訳（大分類単位の配送計画検討用）").font = st['font_title']

    cat_headers = ['店舗名']
    for cat in categories:
        cat_headers += [f'{cat} 出荷', f'{cat} 受入']
    cat_headers += ['合計出荷', '合計受入']

    for ci, hdr in enumerate(cat_headers, sc):
        fill_key = 'fill_navy'
        font_key = 'font_header'
        for cn in categories:
            if cn in hdr:
                fill_key = CAT_FILL.get(cn, 'fill_dark')
                font_key = 'font_header_dk'
                break
        cell = ws.cell(row=cat_title_r, column=ci, value=hdr)
        cell.font      = st[font_key]
        cell.fill      = st[fill_key]
        cell.alignment = st['align_c']
        cell.border    = st['border_thin']
        ws.column_dimensions[get_column_letter(ci)].width = 15

    for ri, store in enumerate(ALL_STORES, cat_title_r+1):
        vals = [store]
        ship_tot = recv_tot = 0
        for cat in categories:
            sq = int(df_all[(df_all['移動元店舗'] == store) & (df_all['大分類'] == cat)]['数量'].sum())
            rq = int(df_all[(df_all['移動先店舗'] == store) & (df_all['大分類'] == cat)]['数量'].sum())
            vals += [sq, rq]
            ship_tot += sq; recv_tot += rq
        vals += [ship_tot, recv_tot]

        for ci, val in enumerate(vals, sc):
            hdr_name = cat_headers[ci - sc]
            fill_key = None
            for cn in categories:
                if cn in hdr_name:
                    fill_key = CAT_FILL.get(cn)
                    break
            is_store = (ci == sc)
            disp = val if is_store or (isinstance(val, int) and val > 0) else '-'
            cell = ws.cell(row=ri, column=ci, value=disp)
            cell.font      = st['font_bold'] if (isinstance(val, int) and val > 0) or is_store else st['font_main']
            cell.alignment = st['align_l'] if is_store else st['align_c']
            cell.border    = st['border_thin']
            if fill_key:
                cell.fill = st[fill_key]

    bot2 = cat_title_r + len(ALL_STORES) + 1
    gs = gr = 0
    tot_vals = ['合計']
    for cat in categories:
        cs = int(df_all[df_all['大分類'] == cat]['数量'].sum())
        tot_vals += [cs, cs]
        gs += cs; gr += cs
    tot_vals += [gs, gr]

    for ci, val in enumerate(tot_vals, sc):
        cell = ws.cell(row=bot2, column=ci, value=val)
        cell.font      = st['font_red_bold']
        cell.fill      = st['fill_yellow']
        cell.alignment = st['align_c']
        cell.border    = st['border_total']


def build_detail_sheet(ws, df_all, st):
    ws.views.sheetView[0].showGridLines = True

    columns = [
        ('大分類',       10, 'c'), ('商品コード', 16, 'c'), ('商品名', 30, 'l'),
        ('カラー', 20, 'l'), ('消化率(%)', 10, 'r'), ('BULK在庫', 10, 'r'),
        ('移動元店舗', 15, 'c'), ('出荷元在庫', 12, 'r'), ('出荷前WOS', 12, 'r'),
        ('移動点数', 12, 'r'),
        ('移動先店舗', 15, 'c'), ('受入先在庫', 12, 'r'), ('受入前WOS', 12, 'r'),
        ('マッチング種別', 18, 'c'),
    ]
    alias = {'移動点数': '数量'}

    for ci, (col_name, col_width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = st['font_header']; cell.fill = st['fill_navy']
        cell.alignment = st['align_c']; cell.border = st['border_thin']
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    df_s = df_all.sort_values(['大分類', '移動元店舗', '移動先店舗', '商品コード'])

    for row_idx, (_, row) in enumerate(df_s.iterrows(), 2):
        cat = str(row.get('大分類', ''))
        bulk_qty = 0
        try:
            bv = row.get('BULK在庫', 0)
            bulk_qty = int(bv) if pd.notna(bv) else 0
        except Exception:
            pass
        row_fill = CAT_FILL.get(cat)

        for ci, (col_name, _, align) in enumerate(columns, 1):
            raw_key = alias.get(col_name, col_name)
            val = row.get(raw_key, '')
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.font = st['font_main']; cell.border = st['border_thin']
            al_map = {'c': 'align_c', 'r': 'align_r', 'l': 'align_l'}
            cell.alignment = st[al_map.get(align, 'align_l')]
            if row_fill:
                cell.fill = st[row_fill]
            if col_name == '移動点数':
                cell.font = st['font_qty_blue']; cell.fill = st['fill_yellow']
            elif col_name in ('移動元店舗', '移動先店舗'):
                cell.font = st['font_bold']
            elif col_name == 'BULK在庫' and bulk_qty > 0:
                cell.fill = st['fill_bulk_org']

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def build_store_sheet(ws, df_all, store_name, st):
    ws.views.sheetView[0].showGridLines = True
    store_df = df_all[df_all['移動元店舗'] == store_name].copy()
    tot_qty  = int(store_df['数量'].sum()) if not store_df.empty else 0

    ws['A1'] = f"出荷指示書： {store_name} → 各受入店舗"
    ws['A1'].font = st['font_title']
    ws['A2'] = f"出荷件数: {len(store_df)} 件 / 合計出荷点数: {tot_qty} 点"
    ws['A2'].font = st['font_subtitle']

    columns = [
        ('大分類', 10, 'c'), ('商品コード', 16, 'c'), ('商品名', 30, 'l'),
        ('カラー', 20, 'l'), ('出荷指示数', 12, 'c'), ('発送先店舗', 15, 'c'),
        ('消化率(%)', 10, 'r'), ('自店現在庫', 12, 'r'), ('自店WOS', 12, 'r'), ('BULK在庫', 10, 'r'),
    ]
    alias = {'出荷指示数': '数量', '発送先店舗': '移動先店舗',
             '自店現在庫': '出荷元在庫', '自店WOS': '出荷前WOS'}

    header_r = 4
    for ci, (col_name, col_width, _) in enumerate(columns, 1):
        cell = ws.cell(row=header_r, column=ci, value=col_name)
        cell.font = st['font_header']; cell.fill = st['fill_blue']
        cell.alignment = st['align_c']; cell.border = st['border_thin']
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    if store_df.empty:
        ws.cell(row=5, column=1, value="この店舗からの出荷対象アイテムはありません。").font = st['font_main']
        return

    store_df = store_df.sort_values(['大分類', '移動先店舗', '商品コード'])
    cur_r = header_r + 1
    ship_total = 0

    for _, row in store_df.iterrows():
        cat = str(row.get('大分類', ''))
        ship_total += int(row.get('数量', 0))
        bulk_qty = 0
        try:
            bv = row.get('BULK在庫', 0)
            bulk_qty = int(bv) if pd.notna(bv) else 0
        except Exception:
            pass
        row_fill = CAT_FILL.get(cat)

        for ci, (col_name, _, align) in enumerate(columns, 1):
            raw_key = alias.get(col_name, col_name)
            val = row.get(raw_key, '')
            cell = ws.cell(row=cur_r, column=ci, value=val)
            cell.font = st['font_main']; cell.border = st['border_thin']
            al_map = {'c': 'align_c', 'r': 'align_r', 'l': 'align_l'}
            cell.alignment = st[al_map.get(align, 'align_l')]
            if row_fill:
                cell.fill = st[row_fill]
            if col_name == '出荷指示数':
                cell.font = st['font_qty_blue']; cell.fill = st['fill_yellow']
            elif col_name == '発送先店舗':
                cell.font = st['font_bold']; cell.fill = st['fill_yellow']
            elif col_name == 'BULK在庫' and bulk_qty > 0:
                cell.fill = st['fill_bulk_org']
        cur_r += 1

    # 合計行
    thin = Side(border_style='thin', color='D9D9D9')
    dbl  = Side(border_style='double', color='000000')
    tot_border = Border(left=thin, right=thin, top=thin, bottom=dbl)
    fill_tot   = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    fill_yel   = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for ci in range(1, len(columns)+1):
        ws.cell(row=cur_r, column=ci).border = tot_border
        ws.cell(row=cur_r, column=ci).fill   = fill_tot
    lc = ws.cell(row=cur_r, column=1, value='合計出荷点数')
    lc.font = st['font_bold']; lc.alignment = st['align_c']
    qc = ws.cell(row=cur_r, column=5, value=ship_total)
    qc.font = st['font_red_bold']; qc.fill = fill_yel; qc.alignment = st['align_c']
    ws.freeze_panes = 'A5'


def main():
    if sys.platform == 'win32':
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    parser = argparse.ArgumentParser(description='暫定移動明細から出荷リストを生成します')
    parser.add_argument('--input',  default='暫定移動明細.xlsx')
    parser.add_argument('--master', default='商品マスタ.csv')
    parser.add_argument('--csv',    default='transfer_upload.csv')
    parser.add_argument('--excel',  default='WOS_Report_店舗出荷.xlsx')
    args = parser.parse_args()

    for path in [args.input, args.master]:
        if not os.path.exists(path):
            print(f"[エラー] ファイルが見つかりません: {path}")
            sys.exit(1)

    print("=== 出荷リスト生成処理 開始 ===")
    records, store_stock, store_wos, item_info = load_and_match(args.input, args.master)
    export_csv(records, args.csv)
    export_excel(records, store_stock, store_wos, item_info, args.excel)

    df_check = pd.DataFrame(records)
    cat_stats = df_check.groupby('大分類')['数量'].sum()
    print("\n=== 完了サマリー ===")
    print(f"総移動点数: {int(df_check['数量'].sum())} 点 ({len(df_check)} 件)")
    for cat, qty in cat_stats.items():
        print(f"  {cat}: {int(qty)} 点")
    print("=" * 30)


if __name__ == '__main__':
    main()
