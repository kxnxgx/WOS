"""
create_shipping_report.py
-------------------------
最新の目星データをもとに、新アルゴリズム（通年WOS補完＋滞留回収＋1便5点おまとめ＋NARITA閉店ルール）
を適用して「WOS_Report_店舗出荷.xlsx」を生成するスクリプト。
"""

import os
import sys
import shutil
import math
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# プロジェクトルートを sys.path に追加して直接実行可能にする
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# ローカルモジュールインポート
from src.data_loader import DataLoader
from src.wos_calculator import WOSCalculator
from src.item_allocator import ItemAllocator

# 店舗名正規化マッピング
STORE_NAME_MAP = {
    'FJALLRAVEN by 3NITY TOKYO': 'TOKYO',
    'TOKYO': 'TOKYO',
    'FJALLRAVEN by 3NITY TOKYO NODE': 'TOKYO NODE',
    'TOKYO NODE': 'TOKYO NODE',
    'FJALLRAVEN by 3NITY 京王新宿': '京王新宿',
    '京王新宿': '京王新宿',
    'FJALLRAVEN by 3NITY玉川高島屋S・C': '玉川高島屋',
    '玉川高島屋': '玉川高島屋',
    'FJALLRAVEN POPUP NARITA': 'NARITA',
    'NARITA': 'NARITA',
    'FJALLRAVEN by 3NITY SAPPORO HUTTE': 'SAPPORO HUTTE',
    'SAPPORO HUTTE': 'SAPPORO HUTTE',
    'FJALLRAVEN STORE 名古屋ファッションワン': '名古屋',
    '名古屋': '名古屋',
    'FJALLRAVEN by 3NITY 大丸心斎橋': '大丸心斎橋',
    '大丸心斎橋': '大丸心斎橋',
    'FJALLRAVEN by 3NITY ルクア大阪': 'ルクア大阪',
    'ルクア大阪': 'ルクア大阪',
}

ALL_STORES = [
    'TOKYO', 'TOKYO NODE', '京王新宿', '玉川高島屋',
    'NARITA', 'SAPPORO HUTTE', '名古屋', '大丸心斎橋', 'ルクア大阪'
]

def normalize_store(name):
    if not name:
        return ''
    s = str(name).strip()
    return STORE_NAME_MAP.get(s, s)

def _is_yellow_fill(fill) -> bool:
    """セルの背景色が黄色系（ハイライト）かどうかを堅牢に判定する"""
    if not fill or not fill.fill_type or fill.fill_type == 'none':
        return False

    # fgColor と start_color の両方をチェック
    for color_prop in [getattr(fill, 'fgColor', None), getattr(fill, 'start_color', None)]:
        if not color_prop:
            continue

        # RGB 判定
        rgb = getattr(color_prop, 'rgb', None)
        if rgb:
            rgb_str = str(rgb).upper()
            if any(y_code in rgb_str for y_code in ('FFFF00', 'FFF2CC', 'FFEAA7', 'FEF08A', 'FDE047')):
                return True

        # Indexed カラー判定（Excel標準の黄色インデックスは 13 または 27）
        if getattr(color_prop, 'type', None) == 'indexed':
            if getattr(color_prop, 'indexed', None) in (13, 27, 34, 43):
                return True

        # テーマカラー値判定
        if getattr(color_prop, 'type', None) == 'theme' and getattr(color_prop, 'value', None) is not None:
            if str(color_prop.value).upper() in ('4', '5', '6', '7') and getattr(color_prop, 'tint', 0) > 0.3:
                return True

    return False

def load_selected_keys_from_excel(filepath):
    """Excelの元ファイルからユーザーが黄色ハイライトした行の (SKU, 出荷店舗, 受入店舗) の組み合わせを読み取る"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    selected_keys = set()

    for sheet_name in ['⭐優先集約リスト', '📦通常集約リスト']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # ヘッダー位置の特定
        sku_col = headers.index('商品コード') + 1 if '商品コード' in headers else 2
        ship_col = headers.index('出荷店舗') + 1 if '出荷店舗' in headers else 7
        recv_col = headers.index('受入店舗') + 1 if '受入店舗' in headers else 12

        for r in range(2, ws.max_row + 1):
            row_cells = ws[r]
            # 黄色ハイライトの判定
            is_yellow = any(_is_yellow_fill(cell.fill) for cell in row_cells)
            
            if is_yellow:
                sku = str(ws.cell(row=r, column=sku_col).value or '').strip()
                ship = normalize_store(ws.cell(row=r, column=ship_col).value)
                recv = normalize_store(ws.cell(row=r, column=recv_col).value)
                if sku and sku != 'None':
                    selected_keys.add((sku, ship, recv))

    return selected_keys

def setup_styles():
    font_main = Font(name='Meiryo UI', size=9)
    font_bold = Font(name='Meiryo UI', size=9, bold=True)
    font_title = Font(name='Meiryo UI', size=13, bold=True, color='1F4E79')
    font_subtitle = Font(name='Meiryo UI', size=9.5, bold=True, color='595959')
    font_header = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFF')

    fill_header_navy = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') 
    fill_header_blue = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid') 
    fill_header_summary = PatternFill(start_color='333F48', end_color='333F48', fill_type='solid') 

    fill_yellow_selected = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
    fill_cont_blue = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid') 
    fill_bulk_orange = PatternFill(start_color='FFEAA7', end_color='FFEAA7', fill_type='solid') 
    fill_highlight_qty = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') 
    fill_total_row = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') 

    thin_border_side = Side(border_style='thin', color='D9D9D9')
    border_all_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom_side = Side(border_style='double', color='000000')
    border_total_row = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_bottom_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    return {
        'font_main': font_main,
        'font_bold': font_bold,
        'font_title': font_title,
        'font_subtitle': font_subtitle,
        'font_header': font_header,
        'fill_header_navy': fill_header_navy,
        'fill_header_blue': fill_header_blue,
        'fill_header_summary': fill_header_summary,
        'fill_yellow_selected': fill_yellow_selected,
        'fill_cont_blue': fill_cont_blue,
        'fill_bulk_orange': fill_bulk_orange,
        'fill_highlight_qty': fill_highlight_qty,
        'fill_total_row': fill_total_row,
        'border_all_thin': border_all_thin,
        'border_total_row': border_total_row,
        'align_center': align_center,
        'align_left': align_left,
        'align_right': align_right,
    }

def calculate_new_moves(sales_df, stock_df, order_df, exclude_stores):
    """通年WOS補完＋滞留回収＋NARITA閉店ルールを組み込んだ新規マッチングロジック"""
    stock_valid = stock_df[~stock_df['store'].apply(lambda s: WOSCalculator.is_excluded_store(s, exclude_stores))].copy()
    sales_valid = sales_df[~sales_df['store'].apply(lambda s: WOSCalculator.is_excluded_store(s, exclude_stores))].copy()

    # 全期間の日数・週数
    min_date = sales_valid['date'].min()
    max_date = sales_valid['date'].max()
    total_weeks = max(1.0, (max_date - min_date).days / 7.0)

    # 1. 通年売上
    annual_sales = sales_valid.groupby(['store', 'sku'])['sales_qty'].sum().reset_index()
    annual_sales['annual_weekly_sales'] = annual_sales['sales_qty'] / total_weeks

    # 2. 直近4週売上
    cutoff_28d = max_date - pd.Timedelta(days=28)
    recent_sales_df = sales_valid[sales_valid['date'] > cutoff_28d]
    recent_sales = recent_sales_df.groupby(['store', 'sku'])['sales_qty'].sum().reset_index()
    recent_sales['recent_weekly_sales'] = recent_sales['sales_qty'] / 4.0

    # 3. マージ
    all_combos = pd.merge(stock_valid[['store', 'sku']], annual_sales[['store', 'sku']], on=['store', 'sku'], how='outer').drop_duplicates()
    wos_new = pd.merge(all_combos, stock_valid, on=['store', 'sku'], how='left')
    wos_new = pd.merge(wos_new, annual_sales[['store', 'sku', 'sales_qty', 'annual_weekly_sales']], on=['store', 'sku'], how='left').rename(columns={'sales_qty': 'annual_sales'})
    wos_new = pd.merge(wos_new, recent_sales[['store', 'sku', 'recent_weekly_sales']], on=['store', 'sku'], how='left')

    wos_new['stock_qty'] = wos_new['stock_qty'].fillna(0)
    wos_new['annual_sales'] = wos_new['annual_sales'].fillna(0)
    wos_new['annual_weekly_sales'] = wos_new['annual_weekly_sales'].fillna(0)
    wos_new['recent_weekly_sales'] = wos_new['recent_weekly_sales'].fillna(0)

    # 消化率
    if order_df is not None and not order_df.empty:
        cum_sales = sales_df.groupby('sku')['sales_qty'].sum().reset_index().rename(columns={'sales_qty': 'cum_sales'})
        ord_clean = order_df.rename(columns={order_df.columns[0]: 'sku', order_df.columns[1]: 'total_order'})
        st_df = pd.merge(cum_sales, ord_clean, on='sku', how='left')
        st_df['sell_through'] = np.where(st_df['total_order'] > 0, (st_df['cum_sales'] / st_df['total_order']) * 100, np.nan)
        wos_new = pd.merge(wos_new, st_df[['sku', 'sell_through']], on='sku', how='left')

    # 商品マスタ
    master = sales_df[['sku', 'item_name', 'color_name']].drop_duplicates('sku')
    wos_new = pd.merge(wos_new, master, on='sku', how='left')

    # 継続品
    wos_new['is_continuation'] = False

    # BULK在庫
    wos_new['bulk_stock'] = 0

    # 実効WOSと週販の算出
    def calc_effective(row):
        stk = row['stock_qty']
        r_spd = row['recent_weekly_sales']
        a_spd = row['annual_weekly_sales']
        if r_spd > 0:
            return pd.Series([r_spd, stk / r_spd, '直近4週'])
        elif a_spd > 0:
            return pd.Series([a_spd, stk / a_spd, '通年補完'])
        elif stk > 0:
            return pd.Series([0.0, 999.0, '完全滞留'])
        else:
            return pd.Series([0.0, np.nan, '在庫なし'])

    wos_new[['effective_speed', 'effective_wos', 'wos_basis']] = wos_new.apply(calc_effective, axis=1)

    move_records = []
    all_skus = wos_new['sku'].unique()

    for sku in all_skus:
        stores = wos_new[wos_new['sku'] == sku]
        valid = stores[(stores['stock_qty'] > 0) | (stores['effective_speed'] > 0)].copy()
        if valid.empty: continue

        active_stores = valid[valid['effective_wos'] < 900]
        wos_avg = active_stores['effective_wos'].mean() if not active_stores.empty else 4.0
        if pd.isna(wos_avg):
            wos_avg = 4.0

        shippers_list = []
        receivers_list = []

        for _, row in valid.iterrows():
            st_name = str(row['store'])
            is_narita = 'NARITA' in st_name
            stk = row['stock_qty']
            r_spd = row['recent_weekly_sales']
            eff_spd = row['effective_speed']
            eff_wos = row['effective_wos']

            if is_narita:
                # NARITA（閉店段階的撤退）
                if stk > 0:
                    keep_qty = math.ceil(r_spd * 2.0) if r_spd > 0 else 0
                    surplus = max(0.0, float(stk - keep_qty))
                    if surplus > 0:
                        r_dict = row.to_dict()
                        r_dict['surplus'] = surplus
                        r_dict['deficit'] = 0.0
                        shippers_list.append(r_dict)
                elif stk == 0 and r_spd >= 0.5:
                    deficit = float(min(2, math.ceil(r_spd * 2.0)))
                    r_dict = row.to_dict()
                    r_dict['surplus'] = 0.0
                    r_dict['deficit'] = deficit
                    receivers_list.append(r_dict)
            else:
                # 通常店舗
                if stk > 0 and (eff_wos > wos_avg or eff_wos >= 900):
                    surplus = float(stk if eff_wos >= 900 else (stk - (wos_avg * eff_spd)))
                    if surplus > 0:
                        r_dict = row.to_dict()
                        r_dict['surplus'] = surplus
                        r_dict['deficit'] = 0.0
                        shippers_list.append(r_dict)
                elif eff_wos < wos_avg and eff_spd > 0:
                    deficit = float((wos_avg * eff_spd) - stk)
                    if deficit > 0:
                        r_dict = row.to_dict()
                        r_dict['surplus'] = 0.0
                        r_dict['deficit'] = deficit
                        receivers_list.append(r_dict)

        if not shippers_list or not receivers_list:
            continue

        shippers_df = pd.DataFrame(shippers_list).sort_values('surplus', ascending=False)
        receivers_df = pd.DataFrame(receivers_list)
        receivers_df['store_rank'] = receivers_df['store'].apply(ItemAllocator._get_store_priority_rank)
        receivers_df = receivers_df.sort_values(by=['deficit', 'store_rank'], ascending=[False, True])

        shippers_dicts = shippers_df.to_dict('records')
        receivers_dicts = receivers_df.to_dict('records')

        st_val = stores['sell_through'].iloc[0] if 'sell_through' in stores.columns else np.nan
        priority = "優先" if pd.notna(st_val) and st_val >= 80.0 else "通常"

        for s in shippers_dicts:
            s['cur_stk'] = s['stock_qty']
        for r in receivers_dicts:
            r['cur_stk'] = r['stock_qty']

        for s in shippers_dicts:
            if s['surplus'] <= 0 or s['cur_stk'] <= 0: continue
            for r in receivers_dicts:
                if r['store'] == s['store']: continue
                if r['deficit'] <= 0: continue
                if s['surplus'] <= 0 or s['cur_stk'] <= 0: break

                move = min(s['surplus'], r['deficit'])
                candidate_qty = max(1, math.ceil(move))
                move_qty = min(candidate_qty, int(s['cur_stk']))

                if move_qty > 0:
                    s['cur_stk'] -= move_qty
                    r['cur_stk'] += move_qty
                    s['surplus'] = max(0.0, s['surplus'] - move_qty)
                    r['deficit'] = max(0.0, r['deficit'] - move_qty)

                    # WOS表記の調整
                    s_pre_wos = f"{s['stock_qty'] / s['recent_weekly_sales']:.1f}週" if s['recent_weekly_sales'] > 0 else "—"
                    s_post_wos = f"{s['cur_stk'] / s['recent_weekly_sales']:.1f}週" if s['recent_weekly_sales'] > 0 else "—"
                    r_pre_wos = f"{r['stock_qty'] / r['recent_weekly_sales']:.1f}週" if r['recent_weekly_sales'] > 0 else "—"
                    r_post_wos = f"{r['cur_stk'] / r['recent_weekly_sales']:.1f}週" if r['recent_weekly_sales'] > 0 else "—"

                    reason = (
                        f"{s['store']}のWOSが{s_pre_wos}（{s['wos_basis']}）、"
                        f"{r['store']}のWOSが{r_pre_wos}（{r['wos_basis']}）のため、"
                        f"{s['store']}から{r['store']}へ{move_qty}点移動を推奨"
                    )

                    move_records.append({
                        'priority': priority,
                        'is_continuation': False,
                        'sku': sku,
                        'item_name': s.get('item_name', ''),
                        'color_name': s.get('color_name', ''),
                        'bulk_stock': 0,
                        'sell_through': st_val,
                        'shipper': s['store'],
                        'shipper_stock': s['stock_qty'],
                        'shipper_pre_wos': s_pre_wos,
                        'shipper_post_wos': s_post_wos,
                        'move_qty': move_qty,
                        'receiver': r['store'],
                        'receiver_stock': r['stock_qty'],
                        'receiver_pre_wos': r_pre_wos,
                        'receiver_post_wos': r_post_wos,
                        'reason': reason
                    })

    df_moves = pd.DataFrame(move_records)
    
    # 店舗名正規化
    df_moves['shipper_norm'] = df_moves['shipper'].apply(normalize_store)
    df_moves['receiver_norm'] = df_moves['receiver'].apply(normalize_store)

    # 1便5点以上のおまとめ最適化
    route_totals = df_moves.groupby(['shipper_norm', 'receiver_norm'])['move_qty'].sum().reset_index().rename(columns={'move_qty': 'route_qty'})
    df_merged = pd.merge(df_moves, route_totals, on=['shipper_norm', 'receiver_norm'], how='left')
    
    # おまとめ適用後のDF
    df_opt = df_merged[df_merged['route_qty'] >= 5].copy()
    df_opt['shipper'] = df_opt['shipper_norm']
    df_opt['receiver'] = df_opt['receiver_norm']
    df_opt = df_opt.drop(columns=['shipper_norm', 'receiver_norm', 'route_qty'])
    
    return df_opt

def build_shipping_summary_sheet(ws, df_selected, styles):
    """シート1: 📋店舗別出荷サマリー の作成"""
    ws.views.sheetView[0].showGridLines = True

    # タイトル
    ws['B2'] = f"【店舗間移動 出荷・受入サマリー（選定アイテム計{len(df_selected)}件）】"
    ws['B2'].font = styles['font_title']
    ws['B3'] = "※各店舗が出荷する点数および受入店舗の内訳一覧です。"
    ws['B3'].font = styles['font_subtitle']

    # 集計マトリクスの作成
    summary_df = df_selected.groupby(['出荷店舗', '受入店舗'])['移動推奨数'].sum().unstack(fill_value=0)
    summary_df = summary_df.reindex(index=ALL_STORES, columns=ALL_STORES, fill_value=0)
    
    start_row = 5
    start_col = 2  # B列から

    # テーブルヘッダー
    ws.cell(row=start_row, column=start_col, value="出荷元店舗 ＼ 受入先店舗").font = styles['font_header']
    ws.cell(row=start_row, column=start_col).fill = styles['fill_header_summary']
    ws.cell(row=start_row, column=start_col).alignment = styles['align_center']
    ws.cell(row=start_row, column=start_col).border = styles['border_all_thin']

    for c_idx, recv_store in enumerate(ALL_STORES, start=start_col + 1):
        cell = ws.cell(row=start_row, column=c_idx, value=recv_store)
        cell.font = styles['font_header']
        cell.fill = styles['fill_header_summary']
        cell.alignment = styles['align_center']
        cell.border = styles['border_all_thin']

    total_col_idx = start_col + len(ALL_STORES) + 1
    cell = ws.cell(row=start_row, column=total_col_idx, value="出荷合計点数")
    cell.font = styles['font_header']
    cell.fill = styles['fill_header_navy']
    cell.alignment = styles['align_center']
    cell.border = styles['border_all_thin']

    item_col_idx = total_col_idx + 1
    cell = ws.cell(row=start_row, column=item_col_idx, value="選定アイテム数")
    cell.font = styles['font_header']
    cell.fill = styles['fill_header_navy']
    cell.alignment = styles['align_center']
    cell.border = styles['border_all_thin']

    # データ行
    for r_idx, ship_store in enumerate(ALL_STORES, start=start_row + 1):
        store_cell = ws.cell(row=r_idx, column=start_col, value=ship_store)
        store_cell.font = styles['font_bold']
        store_cell.fill = styles['fill_total_row']
        store_cell.alignment = styles['align_center']
        store_cell.border = styles['border_all_thin']

        row_total_qty = 0
        for c_idx, recv_store in enumerate(ALL_STORES, start=start_col + 1):
            qty = int(summary_df.loc[ship_store, recv_store])
            row_total_qty += qty
            cell = ws.cell(row=r_idx, column=c_idx, value=qty if qty > 0 else "-")
            cell.font = styles['font_bold'] if qty > 0 else styles['font_main']
            cell.alignment = styles['align_center']
            cell.border = styles['border_all_thin']
            if qty > 0:
                cell.fill = styles['fill_highlight_qty']

        # 出荷合計点数
        cell_total = ws.cell(row=r_idx, column=total_col_idx, value=row_total_qty)
        cell_total.font = styles['font_bold']
        cell_total.alignment = styles['align_center']
        cell_total.border = styles['border_all_thin']
        if row_total_qty > 0:
            cell_total.fill = styles['fill_yellow_selected']

        # 選定アイテム数
        item_count = len(df_selected[df_selected['出荷店舗'] == ship_store])
        cell_item = ws.cell(row=r_idx, column=item_col_idx, value=f"{item_count}件" if item_count > 0 else "-")
        cell_item.font = styles['font_bold'] if item_count > 0 else styles['font_main']
        cell_item.alignment = styles['align_center']
        cell_item.border = styles['border_all_thin']

    # 最下部：受入合計行
    bottom_row = start_row + len(ALL_STORES) + 1
    total_label_cell = ws.cell(row=bottom_row, column=start_col, value="受入合計点数")
    total_label_cell.fill = styles['fill_header_navy']
    total_label_cell.font = styles['font_header']
    total_label_cell.alignment = styles['align_center']
    total_label_cell.border = styles['border_total_row']

    grand_total_qty = 0
    for c_idx, recv_store in enumerate(ALL_STORES, start=start_col + 1):
        col_total_qty = int(summary_df[recv_store].sum())
        grand_total_qty += col_total_qty
        cell = ws.cell(row=bottom_row, column=c_idx, value=col_total_qty if col_total_qty > 0 else "-")
        cell.font = styles['font_bold']
        cell.alignment = styles['align_center']
        cell.border = styles['border_total_row']
        if col_total_qty > 0:
            cell.fill = styles['fill_yellow_selected']

    # 総合計
    grand_total_cell = ws.cell(row=bottom_row, column=total_col_idx, value=grand_total_qty)
    grand_total_cell.font = Font(name='Meiryo UI', size=11, bold=True, color='C00000')
    grand_total_cell.alignment = styles['align_center']
    grand_total_cell.fill = styles['fill_yellow_selected']
    grand_total_cell.border = styles['border_total_row']

    # 総アイテム数
    total_items_cell = ws.cell(row=bottom_row, column=item_col_idx, value=f"{len(df_selected)}件")
    total_items_cell.font = Font(name='Meiryo UI', size=10, bold=True, color='C00000')
    total_items_cell.alignment = styles['align_center']
    total_items_cell.fill = styles['fill_yellow_selected']
    total_items_cell.border = styles['border_total_row']

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 24
    for c_idx in range(start_col + 1, total_col_idx + 2):
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = 16

def build_merged_summary_sheet(ws, df_all, selected_keys, styles):
    """シート2: 📋統合集約リスト の作成"""
    ws.views.sheetView[0].showGridLines = True

    # 優先度順（優先➔通常）＋ 消化率降順
    df_sorted = df_all.copy()
    selected_skus = set(k[0] for k in selected_keys)
    df_sorted['priority_rank'] = df_sorted['優先度'].apply(lambda x: 0 if x == '優先' else 1)
    df_sorted['st_num'] = pd.to_numeric(df_sorted['消化率(%)'], errors='coerce')
    df_sorted = df_sorted.sort_values(by=['priority_rank', 'st_num'], ascending=[True, False]).drop(columns=['priority_rank', 'st_num'])

    columns = [
        ('選定対象', 10, 'center'),
        ('優先度', 10, 'center'),
        ('商品コード', 16, 'center'),
        ('商品名', 28, 'left'),
        ('カラー', 20, 'left'),
        ('BULK在庫', 12, 'right'),
        ('消化率(%)', 12, 'right'),
        ('出荷店舗', 15, 'center'),
        ('出荷元在庫', 12, 'right'),
        ('出荷前WOS', 12, 'right'),
        ('出荷後WOS', 12, 'right'),
        ('移動推奨数', 12, 'right'),
        ('受入店舗', 15, 'center'),
        ('受入先在庫', 12, 'right'),
        ('受入前WOS', 12, 'right'),
        ('受入後WOS', 12, 'right'),
        ('理由', 55, 'left')
    ]

    for col_idx, (col_name, col_width, align) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = styles['font_header']
        cell.fill = styles['fill_header_navy']
        cell.alignment = styles['align_center']
        cell.border = styles['border_all_thin']
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width

    for row_idx, (_, row_data) in enumerate(df_sorted.iterrows(), start=2):
        sku = str(row_data.get('商品コード', '')).strip()
        ship = normalize_store(row_data.get('出荷店舗', ''))
        recv = normalize_store(row_data.get('受入店舗', ''))
        
        # ユーザー選定キーと一致するか判定
        is_sel = sku in selected_skus
        
        item_name = str(row_data.get('商品名', ''))
        is_cont = '🔄' in item_name
        bulk_val = row_data.get('BULK在庫', 0)
        try:
            bulk_qty = int(bulk_val) if pd.notna(bulk_val) else 0
        except:
            bulk_qty = 0

        # 行の基本背景色
        if is_sel:
            row_fill = styles['fill_yellow_selected']
        elif is_cont:
            row_fill = styles['fill_cont_blue']
        else:
            row_fill = None

        for col_idx, (col_name, _, align) in enumerate(columns, start=1):
            if col_name == '選定対象':
                val = '○' if is_sel else ''
            else:
                val = row_data.get(col_name, '')

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles['font_bold'] if is_sel and col_name in ('選定対象', '移動推奨数', '出荷店舗', '受入店舗') else styles['font_main']
            cell.border = styles['border_all_thin']

            if align == 'center':
                cell.alignment = styles['align_center']
            elif align == 'right':
                cell.alignment = styles['align_right']
            else:
                cell.alignment = styles['align_left']

            if row_fill:
                cell.fill = row_fill

            if col_name == 'BULK在庫' and bulk_qty > 0:
                cell.fill = styles['fill_bulk_orange']

            if col_name == '選定対象' and is_sel:
                cell.font = Font(name='Meiryo UI', size=11, bold=True, color='C00000')

    ws.freeze_panes = 'A2'

def build_store_shipping_sheets(wb, df_selected, styles):
    """シート3〜11: 出荷_<店舗名> シートの作成"""
    store_cols = [
        ('商品コード', 16, 'center'),
        ('商品名', 28, 'left'),
        ('カラー', 20, 'left'),
        ('出荷指示数', 12, 'center'),
        ('発送先店舗', 15, 'center'),
        ('優先度', 10, 'center'),
        ('自店現在庫', 12, 'right'),
        ('自店出荷前WOS', 13, 'right'),
        ('自店出荷後WOS', 13, 'right'),
        ('受入先在庫', 12, 'right'),
        ('受入前WOS', 12, 'right'),
        ('受入後WOS', 12, 'right'),
        ('消化率(%)', 12, 'right'),
        ('BULK在庫', 12, 'right'),
        ('移動理由', 55, 'left')
    ]

    col_map_from_raw = {
        '商品コード': '商品コード',
        '商品名': '商品名',
        'カラー': 'カラー',
        '出荷指示数': '移動推奨数',
        '発送先店舗': '受入店舗',
        '優先度': '優先度',
        '自店現在庫': '出荷元在庫',
        '自店出荷前WOS': '出荷前WOS',
        '自店出荷後WOS': '出荷後WOS',
        '受入先在庫': '受入先在庫',
        '受入前WOS': '受入前WOS',
        '受入後WOS': '受入後WOS',
        '消化率(%)': '消化率(%)',
        'BULK在庫': 'BULK在庫',
        '移動理由': '理由'
    }

    for store_name in ALL_STORES:
        store_df = df_selected[df_selected['出荷店舗'] == store_name].copy()
        sheet_title = f"出荷_{store_name}"
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        ws['A1'] = f"【出荷指示書】 {store_name} ➔ 各受入店舗"
        ws['A1'].font = styles['font_title']
        tot_qty_val = int(store_df['移動推奨数'].sum()) if not store_df.empty else 0
        ws['A2'] = f"対象出荷件数: {len(store_df)} 件 / 合計出荷点数: {tot_qty_val} 点"
        ws['A2'].font = styles['font_subtitle']

        header_row = 4
        for col_idx, (col_name, col_width, align) in enumerate(store_cols, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = styles['font_header']
            cell.fill = styles['fill_header_blue']
            cell.alignment = styles['align_center']
            cell.border = styles['border_all_thin']
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_width

        if store_df.empty:
            ws.cell(row=5, column=1, value="※この店舗からの出荷対象アイテムはありません。").font = styles['font_main']
            continue

        store_df['priority_rank'] = store_df['優先度'].apply(lambda x: 0 if x == '優先' else 1)
        store_df = store_df.sort_values(by=['priority_rank', '受入店舗', '商品コード']).drop(columns=['priority_rank'])

        current_row = header_row + 1
        total_ship_qty = 0

        for _, row_data in store_df.iterrows():
            item_name = str(row_data.get('商品名', ''))
            is_cont = '🔄' in item_name
            move_qty = int(row_data.get('移動推奨数', 0)) if pd.notna(row_data.get('移動推奨数')) else 0
            total_ship_qty += move_qty

            bulk_val = row_data.get('BULK在庫', 0)
            try:
                bulk_qty = int(bulk_val) if pd.notna(bulk_val) else 0
            except:
                bulk_qty = 0

            row_fill = styles['fill_cont_blue'] if is_cont else None

            for col_idx, (col_name, _, align) in enumerate(store_cols, start=1):
                raw_field = col_map_from_raw[col_name]
                val = row_data.get(raw_field, '')

                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = styles['font_main']
                cell.border = styles['border_all_thin']

                if row_fill:
                    cell.fill = row_fill

                if align == 'center':
                    cell.alignment = styles['align_center']
                elif align == 'right':
                    cell.alignment = styles['align_right']
                else:
                    cell.alignment = styles['align_left']

                if col_name == '出荷指示数':
                    cell.font = Font(name='Meiryo UI', size=11, bold=True, color='002060')
                    cell.fill = styles['fill_yellow_selected']
                elif col_name == '発送先店舗':
                    cell.font = styles['font_bold']
                    cell.fill = styles['fill_yellow_selected']
                elif col_name == 'BULK在庫' and bulk_qty > 0:
                    cell.fill = styles['fill_bulk_orange']

            current_row += 1

        # 合計行
        tot_label = ws.cell(row=current_row, column=1, value="合計出荷点数")
        tot_label.font = styles['font_bold']
        tot_label.alignment = styles['align_center']
        tot_label.fill = styles['fill_total_row']
        tot_label.border = styles['border_total_row']

        for c in range(2, 4):
            blank_cell = ws.cell(row=current_row, column=c, value="")
            blank_cell.fill = styles['fill_total_row']
            blank_cell.border = styles['border_total_row']

        tot_val = ws.cell(row=current_row, column=4, value=total_ship_qty)
        tot_val.font = Font(name='Meiryo UI', size=11, bold=True, color='C00000')
        tot_val.alignment = styles['align_center']
        tot_val.fill = styles['fill_yellow_selected']
        tot_val.border = styles['border_total_row']

        for c in range(5, len(store_cols) + 1):
            blank_cell = ws.cell(row=current_row, column=c, value="")
            blank_cell.fill = styles['fill_total_row']
            blank_cell.border = styles['border_total_row']

        ws.freeze_panes = 'A5'

def main():
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    source_file = "WOS_Report.xlsx"
    output_file = "WOS_Report_店舗出荷.xlsx"

    print("=== WOS 店舗別出荷指示レポート生成処理 開始 ===")
    print(f"データ読み込み元: {source_file}")

    # 1. ユーザー選定キー（目星の黄色行）を読み込む
    selected_keys = load_selected_keys_from_excel(source_file)
    print(f"ユーザー選定キー数: {len(selected_keys)} 件")

    # 2. 新アルゴリズムによる全提案候補の計算
    sales_csv = '営業日付別売上分析(旧) (18).csv'
    stock_csv = '在庫一覧 (7).csv'
    order_ss = 'FRV26SS発注数.xlsx'
    
    loader = DataLoader(sales_csv, stock_csv)
    sales = loader.load_sales_history()
    stock = loader.load_current_stock()
    order_df = loader.load_order_data(order_ss)

    exclude_stores = [
        '6142', 'FJALLRAVEN POP-UP', 'FJALLRAVEN by 3NITY', 'FJALLRAVEN by 3NITY POPUP',
        'New Way-A', 'New Way-B', 'New Way-C', 'New Way-G', 'ZOZO', 'テスト店舗',
        'バルク', '丸井ｳｪﾌﾞﾁｬﾝﾈﾙ', 'ﾕﾆﾌｫｰﾑ ﾙｸｱ大阪', 'ﾕﾆﾌｫｰﾑNODE', 'ﾕﾆﾌｫｰﾑTOKYO',
        'ﾕﾆﾌｫｰﾑ京王新宿', 'ﾕﾆﾌｫｰﾑ名古屋', 'ﾕﾆﾌｫｰﾑ大丸心斎橋', 'ﾕﾆﾌｫｰﾑ本社', 'ﾕﾆﾌｫｰﾑ玉川高島屋'
    ]

    # 新アルゴリズムの実行（通年WOS補完＋滞留回収＋NARITA閉店段階的撤退＋1便5点おまとめ）
    df_all_moves = calculate_new_moves(sales, stock, order_df, exclude_stores)
    print(f"新アルゴリズム総移動提案候補: {len(df_all_moves)} 件")

    # 日本語カラム名にリネームして統合リスト形式にする
    rename_cols = {
        'priority': '優先度',
        'sku': '商品コード',
        'item_name': '商品名',
        'color_name': 'カラー',
        'bulk_stock': 'BULK在庫',
        'sell_through': '消化率(%)',
        'shipper': '出荷店舗',
        'shipper_stock': '出荷元在庫',
        'shipper_pre_wos': '出荷前WOS',
        'shipper_post_wos': '出荷後WOS',
        'move_qty': '移動推奨数',
        'receiver': '受入店舗',
        'receiver_stock': '受入先在庫',
        'receiver_pre_wos': '受入前WOS',
        'receiver_post_wos': '受入後WOS',
        'reason': '理由'
    }
    df_all_display = df_all_moves.rename(columns=rename_cols)

    # 3. ユーザー選定行（目星行）のみを抽出
    selected_skus = set(k[0] for k in selected_keys)
    
    def is_selected_row(row):
        sku = str(row['商品コード']).strip()
        return sku in selected_skus

    df_selected = df_all_display[df_all_display.apply(is_selected_row, axis=1)].copy()
    print(f"選定された出荷対象行: {len(df_selected)} 件 (総出荷数: {df_selected['移動推奨数'].sum()} 点)")

    # 4. 新規Workbookの構築
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    styles = setup_styles()

    # シート1: 📋店舗別出荷サマリー
    print("生成中: 📋店舗別出荷サマリー シート")
    ws_sum = wb.create_sheet(title='📋店舗別出荷サマリー', index=0)
    build_shipping_summary_sheet(ws_sum, df_selected, styles)

    # シート2: 📋統合集約リスト
    print("生成中: 📋統合集約リスト シート")
    ws_merged = wb.create_sheet(title='📋統合集約リスト', index=1)
    build_merged_summary_sheet(ws_merged, df_all_display, selected_keys, styles)

    # シート3〜11: 出荷_<店舗名>
    print("生成中: 各店舗別出荷シート（全9店舗）")
    build_store_shipping_sheets(wb, df_selected, styles)

    if default_sheet.title in wb.sheetnames:
        wb.remove(default_sheet)

    print(f"Excelファイルを保存しています: {output_file}")
    try:
        wb.save(output_file)
        print("=== レポート生成が正常に完了しました ===")
    except PermissionError:
        import datetime
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        fallback_file = f"WOS_Report_店舗出荷_{timestamp}.xlsx"
        wb.save(fallback_file)
        print(f"\n[⚠️ 警告] '{output_file}' は現在Excel等で開かれているため上書き保存できませんでした。")
        print(f"[✅ 代替保存] '{fallback_file}' として新規保存しました。")
        print(f"※'{output_file}' を更新したい場合は、Excelファイルを閉じてから再度実行してください。")
        print("=== レポート生成が代替ファイル名で完了しました ===")

if __name__ == '__main__':
    main()
