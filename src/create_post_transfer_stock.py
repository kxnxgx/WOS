"""
create_post_transfer_stock.py
------------------------------
暫定移動明細.xlsx の「在庫」+「ORDER」を合算して
集約完了後の想定在庫表を、暫定移動明細と同じフォーマットで出力するスクリプト。

計算ロジック:
  集約後在庫 = 現在庫 + ORDER（ORDER: 正=受入, 負=出荷）
  集約後在庫日数 = 集約後在庫 ÷ (現在庫 ÷ 現在の在庫日数)
                 ※ 現在庫 > 0 かつ 在庫日数 > 0 のときのみ計算
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = r"c:\Users\kesuzuki\Desktop\リテール業務\在庫日数（WOS）"
INPUT_PATH  = os.path.join(WORKSPACE, "暫定移動明細.xlsx")
OUTPUT_PATH = os.path.join(WORKSPACE, "集約後想定在庫.xlsx")

# スタイル
def make_style(fg_color, bold=False, font_color="000000", size=9):
    font = Font(name='Meiryo UI', size=size, bold=bold, color=font_color)
    fill = PatternFill(start_color=fg_color, end_color=fg_color, fill_type='solid') if fg_color else None
    return font, fill

THIN = Side(border_style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_L = Alignment(horizontal='left',   vertical='center')
ALIGN_R = Alignment(horizontal='right',  vertical='center')

def write_cell(ws, row, col, value, fg_color=None, bold=False, font_color="000000",
               size=9, align='center', border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Meiryo UI', size=size, bold=bold, color=font_color)
    if fg_color:
        cell.fill = PatternFill(start_color=fg_color, end_color=fg_color, fill_type='solid')
    if border:
        cell.border = BORDER
    cell.alignment = ALIGN_C if align == 'center' else (ALIGN_R if align == 'right' else ALIGN_L)
    return cell


def main():
    if sys.platform == 'win32':
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    print(f"[1/3] 暫定移動明細.xlsx を読み込んでいます...")
    wb_src = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws_src = wb_src['order'] if 'order' in wb_src.sheetnames else wb_src.active

    row1 = [c.value for c in ws_src[1]]
    row2 = [c.value for c in ws_src[2]]

    # 列インデックスの特定
    sku_col = name_col = color_col = st_col = bulk_col = None
    stock_cols  = []  # [(idx, store_name)]  在庫
    order_cols  = []  # [(idx, store_name)]  ORDER
    wos_cols    = []  # [(idx, store_name)]  在庫日数

    for idx in range(1, len(row1) + 1):
        cat  = str(row1[idx-1] or '').strip()
        item = str(row2[idx-1] or '').strip()
        if item == '商品コード':        sku_col   = idx
        elif item == '商品名':          name_col  = idx
        elif item == 'カラー':          color_col = idx
        elif '消化率' in item:          st_col    = idx
        elif item == 'BULK在庫':        bulk_col  = idx
        elif cat.upper() == '在庫':     stock_cols.append((idx, item))
        elif cat.upper() == 'ORDER':    order_cols.append((idx, item))
        elif cat.upper() == '在庫日数': wos_cols.append((idx, item))

    stores = [sn for _, sn in stock_cols]  # 順序は在庫列に合わせる
    print(f"    対象店舗: {stores}")

    # store_name → インデックスのマップ
    stock_map = {sn: idx for idx, sn in stock_cols}
    order_map = {sn: idx for idx, sn in order_cols}
    wos_map   = {sn: idx for idx, sn in wos_cols}

    # ── データ読み込みと集約後在庫計算 ──
    print("[2/3] 集約後在庫を計算しています...")
    rows = []
    unbalanced_list = []

    for r in range(3, ws_src.max_row + 1):
        sku_val = ws_src.cell(r, sku_col).value
        if not sku_val or str(sku_val).strip() == '':
            continue

        sku          = str(sku_val).strip()
        item_name    = str(ws_src.cell(r, name_col).value  or '').strip() if name_col  else ''
        color_name   = str(ws_src.cell(r, color_col).value or '').strip() if color_col else ''
        sell_through = ws_src.cell(r, st_col).value  if st_col  else None
        bulk_raw     = ws_src.cell(r, bulk_col).value if bulk_col else 0
        bulk_qty     = int(bulk_raw) if isinstance(bulk_raw, (int, float)) else 0

        # 店舗ごとの現在庫・ORDER・在庫日数
        cur_stock  = {}
        order_vals = {}
        cur_wos    = {}

        for sn in stores:
            sv = ws_src.cell(r, stock_map[sn]).value
            ov = ws_src.cell(r, order_map[sn]).value if sn in order_map else None
            wv = ws_src.cell(r, wos_map[sn]).value   if sn in wos_map   else None
            cur_stock[sn]  = int(sv) if isinstance(sv, (int, float)) else 0
            order_vals[sn] = int(ov) if isinstance(ov, (int, float)) else 0
            cur_wos[sn]    = float(wv) if isinstance(wv, (int, float)) and wv is not None else None

        # 不整合チェック
        total_ship = sum(abs(v) for v in order_vals.values() if v < 0)
        total_recv = sum(v for v in order_vals.values() if v > 0)
        if total_ship != total_recv:
            unbalanced_list.append({
                'sku': sku, 'name': item_name, 'row': r,
                'ship': total_ship, 'recv': total_recv, 'diff': total_recv - total_ship
            })

        # 集約後在庫・在庫日数の計算
        post_stock = {}
        post_wos   = {}

        for sn in stores:
            new_stk = cur_stock[sn] + order_vals[sn]
            new_stk = max(0, new_stk)  # マイナス在庫は 0 に丸める
            post_stock[sn] = new_stk

            # 集約後在庫日数 = 新在庫 × (現在の在庫日数 / 現在庫)
            old_wos = cur_wos.get(sn)
            old_stk = cur_stock[sn]
            if old_stk > 0 and old_wos is not None and old_wos > 0:
                daily_speed = old_stk / old_wos  # 1日あたりの販売速度
                new_wos_days = new_stk / daily_speed if daily_speed > 0 else None
                post_wos[sn] = round(new_wos_days, 1) if new_wos_days is not None else None
            elif new_stk > 0:
                # 現在庫 0 で入庫がある場合 → 在庫日数不明
                post_wos[sn] = None
            else:
                post_wos[sn] = None

        rows.append({
            'sku': sku, 'name': item_name, 'color': color_name,
            'sell_through': sell_through, 'bulk': bulk_qty,
            'post_stock': post_stock, 'post_wos': post_wos,
            'order_vals': order_vals,
        })

    print(f"    対象SKU数: {len(rows)}")
    if unbalanced_list:
        print(f"    [警告] 出入庫不一致: {len(unbalanced_list)} SKU")
        for u in unbalanced_list[:5]:
            print(f"      行{u['row']} {u['sku']} {u['name']}: 出庫{u['ship']}点 / 入庫{u['recv']}点 / 差異{u['diff']:+d}点")

    # ── Excel 出力 ──
    print("[3/3] 集約後想定在庫.xlsx を生成しています...")
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = '集約後想定在庫'
    ws.views.sheetView[0].showGridLines = True

    # ヘッダー行1: カテゴリ
    # 列構成: SKU, 商品名, カラー, 消化率(%), BULK在庫 + 在庫×9 + ORDER×9 + 在庫日数×9
    n_stores = len(stores)
    cat_defs = [
        ('商品情報', 5),
        ('在庫', n_stores),
        ('ORDER（参考）', n_stores),
        ('想定在庫日数', n_stores),
    ]

    col = 1
    for cat_name, span in cat_defs:
        if span == 1:
            write_cell(ws, 1, col, cat_name, fg_color='1F4E79', bold=True,
                       font_color='FFFFFF', size=9)
            col += 1
        else:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + span - 1)
            cell = ws.cell(row=1, column=col, value=cat_name)
            cell.font      = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFF')
            cell.alignment = ALIGN_C
            fill_color = {'商品情報': '1F4E79', '在庫': '2E75B6',
                          'ORDER（参考）': '595959', '想定在庫日数': '375623'}.get(cat_name, '333F48')
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.border = BORDER
            for ci in range(col+1, col+span):
                ws.cell(1, ci).border = BORDER
                ws.cell(1, ci).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            col += span

    # ヘッダー行2: 列名
    header2 = ['商品コード', '商品名', 'カラー', '消化率(%)', 'BULK在庫'] + \
              stores + stores + stores

    fill_map = {
        '商品コード': '1F4E79', '商品名': '1F4E79', 'カラー': '1F4E79',
        '消化率(%)': '1F4E79', 'BULK在庫': '1F4E79',
    }
    for ci, hdr in enumerate(header2, 1):
        if ci <= 5:
            fg = '1F4E79'
        elif ci <= 5 + n_stores:
            fg = '2E75B6'
        elif ci <= 5 + 2*n_stores:
            fg = '595959'
        else:
            fg = '375623'
        write_cell(ws, 2, ci, hdr, fg_color=fg, bold=True, font_color='FFFFFF', size=9)

    # データ行
    # 在庫変化の色付け: 増加=薄緑, 減少=薄ピンク, 変化なし=白
    FILL_INC  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 増加（薄緑）
    FILL_DEC  = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # 減少（薄ピンク）
    FILL_ZERO = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 在庫0

    for ri, row_data in enumerate(rows, 3):
        col = 1
        # 商品情報
        write_cell(ws, ri, col,   row_data['sku'],          align='center'); col += 1
        write_cell(ws, ri, col,   row_data['name'],         align='left');   col += 1
        write_cell(ws, ri, col,   row_data['color'],        align='left');   col += 1
        write_cell(ws, ri, col,   row_data['sell_through'], align='right');  col += 1
        write_cell(ws, ri, col,   row_data['bulk'],         align='right');  col += 1

        # 集約後在庫
        for sn in stores:
            new_stk = row_data['post_stock'][sn]
            order_v = row_data['order_vals'][sn]
            cell = ws.cell(row=ri, column=col, value=new_stk if new_stk > 0 else None)
            cell.font      = Font(name='Meiryo UI', size=9,
                                  bold=(new_stk > 0 and order_v != 0))
            cell.alignment = ALIGN_C
            cell.border    = BORDER
            if new_stk == 0:
                cell.fill = FILL_ZERO
            elif order_v > 0:
                cell.fill = FILL_INC  # 受入: 薄緑
            elif order_v < 0:
                cell.fill = FILL_DEC  # 出荷: 薄ピンク
            col += 1

        # ORDER（参考）
        for sn in stores:
            ov = row_data['order_vals'][sn]
            cell = ws.cell(row=ri, column=col, value=ov if ov != 0 else None)
            cell.font      = Font(name='Meiryo UI', size=9,
                                  color='C00000' if ov < 0 else ('0070C0' if ov > 0 else '000000'))
            cell.alignment = ALIGN_C
            cell.border    = BORDER
            if ov < 0:
                cell.fill = FILL_DEC
            elif ov > 0:
                cell.fill = FILL_INC
            col += 1

        # 想定在庫日数
        for sn in stores:
            wos_val = row_data['post_wos'][sn]
            cell = ws.cell(row=ri, column=col, value=round(wos_val, 1) if wos_val is not None else None)
            cell.font      = Font(name='Meiryo UI', size=9)
            cell.alignment = ALIGN_C
            cell.border    = BORDER
            col += 1

    # 列幅設定
    col_widths = [16, 28, 18, 10, 10]  # 商品情報5列
    col_widths += [10] * n_stores       # 在庫
    col_widths += [10] * n_stores       # ORDER
    col_widths += [12] * n_stores       # 在庫日数
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'F3'  # 商品情報5列固定 + ヘッダー2行固定

    # 凡例メモ
    note_row = len(rows) + 4
    ws.cell(note_row, 1, '【色凡例】').font = Font(name='Meiryo UI', size=8, bold=True)
    ws.cell(note_row+1, 1, '薄緑 = 集約受入（在庫増）')
    ws.cell(note_row+1, 1).fill = FILL_INC
    ws.cell(note_row+2, 1, '薄ピンク = 集約出荷（在庫減）')
    ws.cell(note_row+2, 1).fill = FILL_DEC

    wb_out.save(OUTPUT_PATH)
    print(f"    → {OUTPUT_PATH}")

    # サマリー表示
    print("\n=== 集約後在庫 サマリー ===")
    total_pre = total_post = 0
    for row_data in rows:
        for sn in stores:
            pre = row_data['post_stock'][sn] - row_data['order_vals'][sn]
            post = row_data['post_stock'][sn]
            total_pre += max(0, pre)
            total_post += post
    print(f"集約前 全店在庫合計: {total_pre} 点")
    print(f"集約後 全店在庫合計: {total_post} 点")

    if unbalanced_list:
        print(f"\n[注意] {len(unbalanced_list)} SKU に出入庫不一致があります。")
        print("不一致SKU（差異が大きい順）:")
        for u in sorted(unbalanced_list, key=lambda x: abs(x['diff']), reverse=True)[:10]:
            print(f"  {u['sku']}  {u['name']}: 出庫{u['ship']}→入庫{u['recv']} (差異{u['diff']:+d}点)")

    print("=" * 40)


if __name__ == '__main__':
    main()
