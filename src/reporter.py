import pandas as pd
import numpy as np
from datetime import datetime

class Reporter:
    @staticmethod
    def _prepare_move_df(move_df: pd.DataFrame, format_unit: bool = True) -> pd.DataFrame:
        if move_df.empty: return move_df
        
        disp = move_df.copy()

        # 列順序を定義（A〜P列：商品情報 ➔ BULK在庫 ➔ 消化率 ➔ 出荷側 ➔ 移動推奨数 ➔ 受入側）
        desired_cols = [
            'priority', 'sku', 'item_name', 'color_name', 'bulk_stock', 'sell_through',
            'shipper', 'shipper_stock', 'shipper_pre_wos', 'shipper_post_wos',
            'move_qty',
            'receiver', 'receiver_stock', 'receiver_pre_wos', 'receiver_post_wos',
            'reason'
        ]
        cols = [c for c in desired_cols if c in disp.columns]
        disp = disp[cols]

        # WOS値のフォーマット（例: 2.5週）
        if format_unit:
            for wos_col in ['shipper_pre_wos', 'shipper_post_wos', 'receiver_pre_wos', 'receiver_post_wos']:
                if wos_col in disp.columns:
                    disp[wos_col] = disp[wos_col].apply(
                        lambda v: f"{v:.1f}週" if pd.notna(v) else "—"
                    )

        rename_dict = {
            'priority': '優先度',
            'sku': '商品コード',
            'item_name': '商品名',
            'color_name': 'カラー',
            'bulk_stock': 'BULK在庫',
            'sell_through': '消化率(%)',
            'shipper': '出荷店舗',
            'shipper_stock': '出荷元在庫',
            'shipper_pre_wos': '出荷前WOS',
            'shipper_post_wos': '出荷後WOS',
            'move_qty': '移動推奨数',
            'receiver': '受入店舗',
            'receiver_stock': '受入先在庫',
            'receiver_pre_wos': '受入前WOS',
            'receiver_post_wos': '受入後WOS',
            'reason': '理由'
        }
        rename_dict = {k: v for k, v in rename_dict.items() if k in disp.columns}
        return disp.rename(columns=rename_dict)

    @staticmethod
    def _wos_color(val, vmin, vmax):
        """WOS値に応じた背景色を返す（赤=不足、白=平均、青=余剰）"""
        if pd.isna(val) or vmin == vmax:
            return '#ffffff'
        ratio = (val - vmin) / (vmax - vmin)
        if ratio < 0.5:
            t = ratio / 0.5
            r = 255
            g = int(255 * t)
            b = int(255 * t)
        else:
            t = (ratio - 0.5) / 0.5
            r = int(255 * (1 - t))
            g = int(255 * (1 - t))
            b = 255
        return f'rgb({r},{g},{b})'

    @staticmethod
    def _st_color(val):
        """消化率に応じた背景色を返す（緑=高、黄=中、赤=低）"""
        if pd.isna(val):
            return '#f0f0f0'
        if val >= 80:
            return '#d4edda'  # 緑
        elif val >= 60:
            return '#fff3cd'  # 黄
        else:
            return '#fde8e8'  # 赤

    @staticmethod
    def _build_heatmap_table(pivot_df: pd.DataFrame) -> str:
        """WOSヒートマップHTMLテーブルを生成する"""
        all_vals = pivot_df.values.flatten()
        valid_vals = all_vals[~np.isnan(all_vals.astype(float))]
        vmin = float(valid_vals.min()) if len(valid_vals) > 0 else 0
        vmax = float(valid_vals.max()) if len(valid_vals) > 0 else 1

        rows = []
        index_names = [n if n else '' for n in (pivot_df.index.names if hasattr(pivot_df.index, 'names') else [pivot_df.index.name])]
        col_headers = list(pivot_df.columns)
        header_html = ''.join(f'<th>{n}</th>' for n in index_names)
        header_html += ''.join(f'<th>{c}</th>' for c in col_headers)
        rows.append(f'<tr>{header_html}</tr>')

        for idx, row in pivot_df.iterrows():
            idx_vals = idx if isinstance(idx, tuple) else (idx,)
            cells = ''.join(f'<td class="idx-cell">{v}</td>' for v in idx_vals)
            for val in row:
                if pd.isna(val):
                    cells += '<td class="wos-cell na-cell">—</td>'
                else:
                    color = Reporter._wos_color(val, vmin, vmax)
                    cells += f'<td class="wos-cell" style="background:{color};" title="WOS: {val:.1f}週">{val:.1f}</td>'
            rows.append(f'<tr>{cells}</tr>')

        return f'<table class="heatmap-table"><thead>{rows[0]}</thead><tbody>{"".join(rows[1:])}</tbody></table>'

    @staticmethod
    def _build_move_table(df: pd.DataFrame) -> str:
        """移動推奨テーブルHTMLを生成する"""
        if df.empty:
            return '<p class="no-data">該当する移動推奨アイテムはありません</p>'
        
        disp = df.copy()
        
        # 継続品バッジを商品名に追加
        if 'is_continuation' in disp.columns:
            disp['item_name'] = disp.apply(
                lambda row: f"{row['item_name']} <span class='cont-badge'>🔄 継続品</span>" 
                            if row['is_continuation'] else row['item_name'], 
                axis=1
            )
            # is_continuation列は表示不要なので退避
            is_cont_series = disp['is_continuation']
            disp = disp.drop(columns=['is_continuation'])
        else:
            is_cont_series = pd.Series([False] * len(disp), index=disp.index)

        # フォールバックWOS使用フラグを退避（表示用）
        if 'is_shipper_fallback' in disp.columns and 'is_receiver_fallback' in disp.columns:
            is_fallback_series = disp['is_shipper_fallback'] | disp['is_receiver_fallback']
            disp = disp.drop(columns=['is_shipper_fallback', 'is_receiver_fallback'])
        else:
            is_fallback_series = pd.Series([False] * len(disp), index=disp.index)

        # BULK在庫のアラート装飾（>0 の場合は黄色ハイライト）
        if 'bulk_stock' in disp.columns:
            disp['bulk_stock'] = disp['bulk_stock'].apply(
                lambda v: f"<span class='bulk-alert'>{v}</span>" if pd.notna(v) and v > 0 else f"<span class='bulk-zero'>{v}</span>"
            )

        disp = Reporter._prepare_move_df(disp)
        
        # HTMLテーブルを生成
        html = disp.to_html(index=False, classes='move-table', border=0, escape=False)
        
        # 行ごとに継続品クラス・フォールバッククラスを付与
        html_lines = html.splitlines()
        tr_idx = 0
        for i, line in enumerate(html_lines):
            if '<tr>' in line and '<th>' not in line:
                classes = []
                if tr_idx < len(is_cont_series) and is_cont_series.iloc[tr_idx]:
                    classes.append('continuation-row')
                if tr_idx < len(is_fallback_series) and is_fallback_series.iloc[tr_idx]:
                    classes.append('fallback-row')
                if classes:
                    html_lines[i] = line.replace('<tr>', f'<tr class="{" ".join(classes)}">')
                tr_idx += 1
                
        return '\n'.join(html_lines)

    @staticmethod
    def _build_summary_table(wos_df: pd.DataFrame, move_df: pd.DataFrame) -> str:
        """全社サマリーテーブルHTMLを生成する（消化率列・消化率順ソート対応）"""
        if wos_df.empty:
            return '<p>データがありません</p>'

        valid_wos = wos_df[wos_df['wos'].notna()]
        summary = valid_wos.groupby('sku').agg(
            avg_wos=('wos', 'mean'),
            min_wos=('wos', 'min'),
            max_wos=('wos', 'max'),
            store_count=('store', 'nunique'),
        ).reset_index()
        summary['avg_wos'] = summary['avg_wos'].round(1)
        summary['min_wos'] = summary['min_wos'].round(1)
        summary['max_wos'] = summary['max_wos'].round(1)

        if 'item_name' in wos_df.columns:
            master = wos_df[['sku', 'item_name']].drop_duplicates('sku')
            summary = summary.merge(master, on='sku', how='left')
        if 'color_name' in wos_df.columns:
            master_c = wos_df[['sku', 'color_name']].drop_duplicates('sku')
            summary = summary.merge(master_c, on='sku', how='left')

        # BULK在庫列の追加
        if 'bulk_stock' in wos_df.columns:
            bulk_master = wos_df[['sku', 'bulk_stock']].drop_duplicates('sku')
            summary = summary.merge(bulk_master, on='sku', how='left')
            summary['bulk_stock'] = summary['bulk_stock'].fillna(0).astype(int)

        # 消化率列の追加
        has_st = 'sell_through' in wos_df.columns
        if has_st:
            st_master = wos_df[['sku', 'sell_through', 'cumulative_sales', 'total_order']].drop_duplicates('sku')
            summary = summary.merge(st_master, on='sku', how='left')

        # 移動推奨件数
        if not move_df.empty:
            move_counts = move_df.groupby('sku').size().reset_index(name='move_count')
            summary = summary.merge(move_counts, on='sku', how='left')
        else:
            summary['move_count'] = 0
        summary['move_count'] = summary['move_count'].fillna(0).astype(int)

        # カラム整理
        col_order = ['sku']
        if 'item_name' in summary.columns: col_order.append('item_name')
        if 'color_name' in summary.columns: col_order.append('color_name')
        if 'bulk_stock' in summary.columns: col_order.append('bulk_stock')
        if has_st:
            col_order += ['sell_through', 'cumulative_sales', 'total_order']
        col_order += ['store_count', 'avg_wos', 'min_wos', 'max_wos', 'move_count']
        summary = summary[[c for c in col_order if c in summary.columns]]

        # 消化率がある場合は消化率降順、ない場合は平均WOS昇順
        if has_st and 'sell_through' in summary.columns:
            summary = summary.sort_values('sell_through', ascending=False)
        else:
            summary = summary.sort_values('avg_wos', ascending=True)

        col_rename = {
            'sku': '商品コード', 'item_name': '商品名', 'color_name': 'カラー',
            'bulk_stock': 'BULK在庫',
            'sell_through': '消化率(%)', 'cumulative_sales': '累計売上数',
            'total_order': '発注数', 'store_count': '対象店舗数',
            'avg_wos': '平均WOS', 'min_wos': '最小WOS', 'max_wos': '最大WOS',
            'move_count': '移動推奨件数'
        }
        summary = summary.rename(columns={k: v for k, v in col_rename.items() if k in summary.columns})

        overall_avg_wos = summary['平均WOS'].mean()

        rows = []
        headers = list(summary.columns)
        rows.append('<tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr>')

        for _, row in summary.iterrows():
            cells = ''
            for col in headers:
                val = row[col]
                cell_style = ''
                if col == '消化率(%)' and pd.notna(val):
                    bg = Reporter._st_color(val)
                    cell_style = f' style="background:{bg};font-weight:bold;"'
                    val = f'{val:.1f}%'
                elif col == '移動推奨件数' and val > 0:
                    cell_style = ' style="background:#fff3cd;font-weight:bold;"'
                elif col == '平均WOS' and pd.notna(val):
                    if val < overall_avg_wos * 0.7:
                        cell_style = ' style="background:#fde8e8;"'
                    elif val > overall_avg_wos * 1.3:
                        cell_style = ' style="background:#e8f0fd;"'
                cells += f'<td{cell_style}>{val}</td>'
            rows.append(f'<tr>{cells}</tr>')

        return (
            f'<p style="color:#666;font-size:0.9em;">全商品の平均WOS: <strong>{overall_avg_wos:.1f}週</strong></p>'
            f'<table class="summary-table"><thead>{rows[0]}</thead><tbody>{"".join(rows[1:])}</tbody></table>'
        )

    @staticmethod
    def _apply_excel_formatting(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Excelのシートに書式（継続品の色付けなど）を適用する"""
        from openpyxl.styles import PatternFill
        workbook = writer.book
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # 青系の薄い色（継続品）
            cont_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            # 薄いオレンジ（フォールバックWOS: 全期間平均週販使用）
            fallback_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            # 薄いオレンジ/黄色（BULK在庫注意）
            bulk_fill = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")

            # フォールバックWOS使用行を先に色付け（後で継続品が上書きするので順序注意）
            has_s_fallback = 'is_shipper_fallback' in df.columns
            has_r_fallback = 'is_receiver_fallback' in df.columns
            if has_s_fallback or has_r_fallback:
                for row_idx in range(len(df)):
                    s_fb = df['is_shipper_fallback'].iloc[row_idx] if has_s_fallback else False
                    r_fb = df['is_receiver_fallback'].iloc[row_idx] if has_r_fallback else False
                    if s_fb or r_fb:
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx + 2, column=col_idx).fill = fallback_fill

            # is_continuation はデータフレームのどこにあるか？
            if 'is_continuation' in df.columns:
                is_cont_series = df['is_continuation'].reset_index(drop=True)
                for row_idx, is_cont in enumerate(is_cont_series):
                    if is_cont:
                        # ヘッダーが1行目なのでデータは2行目から
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx + 2, column=col_idx).fill = cont_fill

            # BULK在庫 のハイライト（列位置を動的に特定）
            if 'bulk_stock' in df.columns:
                bulk_series = df['bulk_stock'].reset_index(drop=True)
                headers = [cell.value for cell in worksheet[1]]
                bulk_col_idx = headers.index('BULK在庫') + 1 if 'BULK在庫' in headers else 5
                for row_idx, b_val in enumerate(bulk_series):
                    if pd.notna(b_val) and b_val > 0:
                        worksheet.cell(row=row_idx + 2, column=bulk_col_idx).fill = bulk_fill

    @staticmethod
    def build_all_stock_df(wos_df: pd.DataFrame, move_df: pd.DataFrame) -> pd.DataFrame:
        """
        全SKU×店舗の在庫データ（wos_df）に消化率および集約先・集約元の移動推奨情報を結合したDataFrameを構築する
        """
        if wos_df.empty:
            return pd.DataFrame()

        df_base = wos_df.copy()

        # 出荷側・受入側の移動推奨情報の集約
        def _format_stores(group, store_col, qty_col):
            if len(group) == 1:
                st = str(group[store_col].iloc[0])
                q = int(group[qty_col].iloc[0])
                return f"{st} ({q}点)"
            parts = [f"{row[store_col]} ({int(row[qty_col])}点)" for _, row in group.iterrows()]
            return ', '.join(parts)

        if not move_df.empty:
            ship_grouped = move_df.groupby(['shipper', 'sku'])
            ship_info = ship_grouped.apply(lambda g: pd.Series({
                'ship_to_stores': _format_stores(g, 'receiver', 'move_qty'),
                'ship_qty': int(g['move_qty'].sum()),
                'ship_priority': '⭐優先' if '優先' in g['priority'].values else '📦通常',
                'ship_reason': '; '.join(g['reason'].dropna().unique())
            }), include_groups=False).reset_index()

            recv_grouped = move_df.groupby(['receiver', 'sku'])
            recv_info = recv_grouped.apply(lambda g: pd.Series({
                'recv_from_stores': _format_stores(g, 'shipper', 'move_qty'),
                'recv_qty': int(g['move_qty'].sum()),
                'recv_priority': '⭐優先' if '優先' in g['priority'].values else '📦通常',
                'recv_reason': '; '.join(g['reason'].dropna().unique())
            }), include_groups=False).reset_index()

            df_merged = pd.merge(df_base, ship_info, left_on=['store', 'sku'], right_on=['shipper', 'sku'], how='left')
            df_merged = pd.merge(df_merged, recv_info, left_on=['store', 'sku'], right_on=['receiver', 'sku'], how='left')
        else:
            df_merged = df_base.copy()
            df_merged['ship_to_stores'] = ''
            df_merged['ship_qty'] = np.nan
            df_merged['ship_priority'] = ''
            df_merged['ship_reason'] = ''
            df_merged['recv_from_stores'] = ''
            df_merged['recv_qty'] = np.nan
            df_merged['recv_priority'] = ''
            df_merged['recv_reason'] = ''

        # 優先度と移動理由の統合
        def _get_combined_priority(row):
            sp = row.get('ship_priority')
            rp = row.get('recv_priority')
            if sp == '⭐優先' or rp == '⭐優先':
                return '⭐優先'
            elif sp == '📦通常' or rp == '📦通常':
                return '📦通常'
            return ''

        def _get_combined_reason(row):
            sr = str(row.get('ship_reason', '') or '')
            rr = str(row.get('recv_reason', '') or '')
            reasons = [r for r in [sr, rr] if r and r != 'nan']
            return ' / '.join(reasons) if reasons else ''

        df_merged['priority'] = df_merged.apply(_get_combined_priority, axis=1)
        df_merged['move_reason'] = df_merged.apply(_get_combined_reason, axis=1)

        # 継続品表記
        is_cont = df_merged['is_continuation'] if 'is_continuation' in df_merged.columns else pd.Series([False] * len(df_merged))
        df_merged['cont_mark'] = is_cont.apply(lambda x: '○' if x is True else '')

        # 列の整理と日本語ヘッダーへのマッピング
        col_mapping = [
            ('store', '店舗名'),
            ('sku', '商品コード'),
            ('item_name', '商品名'),
            ('color_name', 'カラー'),
            ('stock_qty', '現在庫'),
            ('avg_sales_4w', '直近4週平均週販'),
            ('wos', 'WOS(週)'),
            ('avg_sales_full', '全期間平均週販'),
            ('wos_fallback', '参考WOS(週)'),
            ('sell_through', '消化率(%)'),
            ('cumulative_sales', '累計売上数'),
            ('total_order', '発注数'),
            ('bulk_stock', 'BULK在庫'),
            ('cont_mark', '継続品'),
            ('ship_to_stores', '集約先店舗（発送先）'),
            ('ship_qty', '集約出荷数'),
            ('recv_from_stores', '集約元店舗（入荷元）'),
            ('recv_qty', '集約受入数'),
            ('priority', '優先度'),
            ('move_reason', '移動理由'),
        ]

        result_df = pd.DataFrame()
        for orig_col, new_col in col_mapping:
            if orig_col in df_merged.columns:
                result_df[new_col] = df_merged[orig_col]
            else:
                result_df[new_col] = np.nan

        # 数値フォーマットの整形
        if '現在庫' in result_df.columns:
            result_df['現在庫'] = pd.to_numeric(result_df['現在庫'], errors='coerce').fillna(0).astype(int)
        if '直近4週平均週販' in result_df.columns:
            result_df['直近4週平均週販'] = pd.to_numeric(result_df['直近4週平均週販'], errors='coerce').round(2)
        if 'WOS(週)' in result_df.columns:
            result_df['WOS(週)'] = pd.to_numeric(result_df['WOS(週)'], errors='coerce').round(1)
        if '全期間平均週販' in result_df.columns:
            result_df['全期間平均週販'] = pd.to_numeric(result_df['全期間平均週販'], errors='coerce').round(2)
        if '参考WOS(週)' in result_df.columns:
            result_df['参考WOS(週)'] = pd.to_numeric(result_df['参考WOS(週)'], errors='coerce').round(1)
        if '消化率(%)' in result_df.columns:
            result_df['消化率(%)'] = pd.to_numeric(result_df['消化率(%)'], errors='coerce').round(1)
        if '累計売上数' in result_df.columns:
            result_df['累計売上数'] = pd.to_numeric(result_df['累計売上数'], errors='coerce').fillna(0).astype(int)
        if '発注数' in result_df.columns:
            result_df['発注数'] = pd.to_numeric(result_df['発注数'], errors='coerce').fillna(0).astype(int)
        if 'BULK在庫' in result_df.columns:
            result_df['BULK在庫'] = pd.to_numeric(result_df['BULK在庫'], errors='coerce').fillna(0).astype(int)

        return result_df

    @staticmethod
    def _apply_all_stock_formatting(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """全店現在庫一覧シートの装飾とスタイリング"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        workbook = writer.book
        if sheet_name not in workbook.sheetnames:
            return

        ws = workbook[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = 'A2'

        font_header = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFF')
        font_main = Font(name='Meiryo UI', size=9)
        font_bold = Font(name='Meiryo UI', size=9, bold=True)

        fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        fill_ship = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 薄い黄色
        fill_recv = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 薄い緑色
        fill_cont = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')  # 薄い青
        fill_bulk = PatternFill(start_color='FFEAA7', end_color='FFEAA7', fill_type='solid')  # 薄いオレンジ

        thin_side = Side(border_style='thin', color='D9D9D9')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        headers = [cell.value for cell in ws[1]]

        # ヘッダー書式設定
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        # データ行書式設定
        center_cols = {'店舗名', '商品コード', '継続品', '優先度'}
        right_cols = {'現在庫', '直近4週平均週販', 'WOS(週)', '全期間平均週販', '参考WOS(週)',
                      '消化率(%)', '累計売上数', '発注数', 'BULK在庫', '集約出荷数', '集約受入数'}

        ship_to_idx = headers.index('集約先店舗（発送先）') + 1 if '集約先店舗（発送先）' in headers else -1
        ship_qty_idx = headers.index('集約出荷数') + 1 if '集約出荷数' in headers else -1
        recv_from_idx = headers.index('集約元店舗（入荷元）') + 1 if '集約元店舗（入荷元）' in headers else -1
        recv_qty_idx = headers.index('集約受入数') + 1 if '集約受入数' in headers else -1
        bulk_idx = headers.index('BULK在庫') + 1 if 'BULK在庫' in headers else -1
        cont_idx = headers.index('継続品') + 1 if '継続品' in headers else -1

        for r_idx in range(2, ws.max_row + 1):
            for c_idx, h_name in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = font_main
                cell.border = border_all

                if h_name in center_cols:
                    cell.alignment = align_center
                elif h_name in right_cols:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

            # 個別ハイライト
            if ship_qty_idx > 0 and pd.notna(ws.cell(row=r_idx, column=ship_qty_idx).value):
                ws.cell(row=r_idx, column=ship_qty_idx).fill = fill_ship
                ws.cell(row=r_idx, column=ship_qty_idx).font = font_bold
                if ship_to_idx > 0:
                    ws.cell(row=r_idx, column=ship_to_idx).fill = fill_ship

            if recv_qty_idx > 0 and pd.notna(ws.cell(row=r_idx, column=recv_qty_idx).value):
                ws.cell(row=r_idx, column=recv_qty_idx).fill = fill_recv
                ws.cell(row=r_idx, column=recv_qty_idx).font = font_bold
                if recv_from_idx > 0:
                    ws.cell(row=r_idx, column=recv_from_idx).fill = fill_recv

            if bulk_idx > 0:
                b_val = ws.cell(row=r_idx, column=bulk_idx).value
                if pd.notna(b_val) and isinstance(b_val, (int, float)) and b_val > 0:
                    ws.cell(row=r_idx, column=bulk_idx).fill = fill_bulk

            if cont_idx > 0 and ws.cell(row=r_idx, column=cont_idx).value == '○':
                ws.cell(row=r_idx, column=cont_idx).fill = fill_cont

        # 列幅の自動調整
        for c_idx, h_name in enumerate(headers, start=1):
            col_letter = get_column_letter(c_idx)
            max_len = len(str(h_name or '')) * 2
            if h_name in ('商品名', '移動理由'):
                ws.column_dimensions[col_letter].width = 35
            elif h_name in ('集約先店舗（発送先）', '集約元店舗（入荷元）'):
                ws.column_dimensions[col_letter].width = 28
            elif h_name in ('店舗名', 'カラー'):
                ws.column_dimensions[col_letter].width = 20
            elif h_name == '商品コード':
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @staticmethod
    def generate_excel(wos_df: pd.DataFrame, move_df: pd.DataFrame, output_path: str = "WOS_Report.xlsx"):
        """Excelレポートの出力（既存4シート ＋ 📋全店現在庫一覧）"""
        print(f"Excelレポートを生成しています: {output_path}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 優先集約リスト
            priority_df = move_df[move_df['priority'] == '優先'].copy() if not move_df.empty and 'priority' in move_df.columns else pd.DataFrame()
            normal_df = move_df[move_df['priority'] != '優先'].copy() if not move_df.empty and 'priority' in move_df.columns else move_df.copy()

            # 表示用DF作成時に is_continuation は削除されるため、元のDataFrameを保持
            def _prepare_and_drop(df):
                if df.empty: return df
                disp = df.copy()
                if 'is_continuation' in disp.columns:
                    disp['item_name'] = disp.apply(
                        lambda row: f"🔄 {row['item_name']}" if row['is_continuation'] else row['item_name'], axis=1
                    )
                    disp = disp.drop(columns=['is_continuation'])
                return Reporter._prepare_move_df(disp)

            disp_priority = _prepare_and_drop(priority_df)
            disp_normal = _prepare_and_drop(normal_df)

            if not disp_priority.empty:
                disp_priority.to_excel(writer, sheet_name='⭐優先集約リスト', index=False)
                Reporter._apply_excel_formatting(writer, '⭐優先集約リスト', priority_df)
            else:
                pd.DataFrame([{"Message": "優先集約アイテムはありません"}]).to_excel(writer, sheet_name='⭐優先集約リスト', index=False)

            if not disp_normal.empty:
                disp_normal.to_excel(writer, sheet_name='📦通常集約リスト', index=False)
                Reporter._apply_excel_formatting(writer, '📦通常集約リスト', normal_df)
            else:
                pd.DataFrame([{"Message": "通常集約アイテムはありません"}]).to_excel(writer, sheet_name='📦通常集約リスト', index=False)

            if not wos_df.empty:
                idx_cols = ['sku']
                if 'item_name' in wos_df.columns: idx_cols.append('item_name')
                if 'color_name' in wos_df.columns: idx_cols.append('color_name')
                if 'sell_through' in wos_df.columns: idx_cols.append('sell_through')
                if 'is_continuation' in wos_df.columns: idx_cols.append('is_continuation')

                pivot_wos = wos_df.pivot(index=idx_cols, columns='store', values='wos').round(1)
                pivot_wos.to_excel(writer, sheet_name='店舗別WOSサマリー')
                wos_df.to_excel(writer, sheet_name='WOS生データ', index=False)

                # シート5: 📋全店現在庫一覧（新シート追加）
                all_stock_df = Reporter.build_all_stock_df(wos_df, move_df)
                if not all_stock_df.empty:
                    all_stock_df.to_excel(writer, sheet_name='📋全店現在庫一覧', index=False)
                    Reporter._apply_all_stock_formatting(writer, '📋全店現在庫一覧', all_stock_df)

    @staticmethod
    def generate_all_stock_excel(wos_df: pd.DataFrame, move_df: pd.DataFrame, output_path: str = "WOS_全店在庫一覧.xlsx"):
        """コピー用別ファイル（WOS_全店在庫一覧.xlsx）の出力"""
        print(f"全店在庫一覧Excelを生成しています: {output_path}")
        all_stock_df = Reporter.build_all_stock_df(wos_df, move_df)
        if all_stock_df.empty:
            print("  [警告] 全店在庫データが空のため、ファイル生成をスキップしました。")
            return

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            sheet_name = '全店現在庫一覧'
            all_stock_df.to_excel(writer, sheet_name=sheet_name, index=False)
            Reporter._apply_all_stock_formatting(writer, sheet_name, all_stock_df)
        print(f"  -> {output_path} 出力完了 ({len(all_stock_df)} 行)")

    @staticmethod
    def generate_html(wos_df: pd.DataFrame, move_df: pd.DataFrame,
                      output_path: str = "WOS_Report.html",
                      threshold: float = 80.0):
        """HTMLレポートの出力（4タブ・消化率対応）"""
        print(f"HTMLレポートを生成しています: {output_path}")

        generated_at = datetime.now().strftime('%Y年%m月%d日 %H:%M')
        has_sell_through = not move_df.empty and 'sell_through' in move_df.columns

        # --- 優先集約 / 通常集約 の分離 ---
        if not move_df.empty and 'priority' in move_df.columns:
            priority_move = move_df[move_df['priority'] == '優先'].copy()
            normal_move = move_df[move_df['priority'] != '優先'].copy()
        else:
            priority_move = pd.DataFrame()
            normal_move = move_df.copy()

        # 消化率降順ソート
        if not priority_move.empty and 'sell_through' in priority_move.columns:
            priority_move = priority_move.sort_values('sell_through', ascending=False)
        if not normal_move.empty and 'sell_through' in normal_move.columns:
            normal_move = normal_move.sort_values('sell_through', ascending=False)

        priority_html = Reporter._build_move_table(priority_move)
        normal_html = Reporter._build_move_table(normal_move)
        summary_html = Reporter._build_summary_table(wos_df, move_df)

        # WOS ヒートマップ
        if not wos_df.empty:
            idx_cols = ['sku']
            if 'item_name' in wos_df.columns: idx_cols.append('item_name')
            if 'color_name' in wos_df.columns: idx_cols.append('color_name')
            pivot_wos = wos_df.pivot(index=idx_cols, columns='store', values='wos').round(1)
            heatmap_html = Reporter._build_heatmap_table(pivot_wos)
        else:
            heatmap_html = '<p class="no-data">データがありません</p>'

        priority_count = len(priority_move)
        normal_count = len(normal_move)
        sku_count = wos_df['sku'].nunique() if not wos_df.empty else 0
        store_count = wos_df['store'].nunique() if not wos_df.empty else 0

        st_badge = f'（消化率 ≥ {threshold:.0f}% 基準）' if has_sell_through else '（消化率データなし）'

        html_content = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WOS アイテム集約レポート</title>
    <style>
        *, *::before, *::after {{ box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', 'Meiryo', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0; padding: 20px 30px; color: #2d3748;
            background: #f7fafc;
        }}
        h1 {{
            color: #1a202c; font-size: 1.6em;
            border-bottom: 3px solid #4299e1; padding-bottom: 10px;
            margin-bottom: 6px;
        }}
        .meta {{ color: #718096; font-size: 0.85em; margin-bottom: 20px; }}
        .kpi-bar {{ display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }}
        .kpi {{
            background: #fff; border-radius: 8px; padding: 14px 22px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.1); min-width: 140px;
        }}
        .kpi .label {{ font-size: 0.78em; color: #718096; margin-bottom: 4px; }}
        .kpi .value {{ font-size: 1.7em; font-weight: 700; color: #2b6cb0; }}
        .kpi.priority .value {{ color: #c05621; }}
        .logic-box {{
            background: #ebf8ff; border-left: 4px solid #4299e1;
            padding: 10px 16px; margin-bottom: 20px; border-radius: 0 6px 6px 0;
            font-size: 0.88em;
        }}
        .logic-box summary {{ font-weight: bold; cursor: pointer; color: #2b6cb0; }}
        .logic-box ul {{ margin: 8px 0 0 0; padding-left: 20px; color: #4a5568; }}
        .tabs {{ display: flex; border-bottom: 2px solid #e2e8f0; margin-bottom: 20px; }}
        .tab {{
            padding: 10px 24px; cursor: pointer; border: none;
            background: none; font-size: 0.95em; color: #718096;
            border-bottom: 3px solid transparent; margin-bottom: -2px;
            transition: color 0.15s;
        }}
        .tab:hover {{ color: #4299e1; }}
        .tab.active {{ color: #2b6cb0; font-weight: bold; border-bottom-color: #4299e1; }}
        .tab.priority-tab.active {{ color: #c05621; border-bottom-color: #ed8936; }}
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
        /* 移動推奨テーブル */
        .move-table {{
            border-collapse: collapse; width: 100%;
            box-shadow: 0 1px 6px rgba(0,0,0,0.08); border-radius: 8px; overflow: hidden;
        }}
        .move-table th {{
            background: #2b6cb0; color: #fff;
            padding: 10px 14px; text-align: left; font-weight: 600; font-size: 0.88em;
        }}
        .move-table td {{
            border-bottom: 1px solid #e2e8f0; padding: 9px 14px; font-size: 0.87em;
        }}
        .move-table tr:last-child td {{ border-bottom: none; }}
        .move-table tr:hover td {{ background: #ebf8ff; }}
        .move-table tr.continuation-row td {{ background: #e6f2ff; }}
        .move-table tr.continuation-row:hover td {{ background: #cce6ff; }}
        .move-table tr.fallback-row td {{ background: #fff8f0; border-left: 3px solid #ed8936; }}
        .move-table tr.fallback-row:hover td {{ background: #feebc8; }}
        .move-table tr.continuation-row.fallback-row td {{ background: #e6f2ff; border-left: 3px solid #ed8936; }}
        .cont-badge {{
            display: inline-block; background: #e1effe; color: #1e40af;
            border: 1px solid #93c5fd; border-radius: 4px;
            padding: 1px 6px; font-size: 0.82em; margin-left: 6px;
        }}
        .bulk-alert {{
            display: inline-block; background: #ffeaa7; color: #b7791f;
            border: 1px solid #fdcb6e; border-radius: 4px;
            padding: 1px 8px; font-weight: bold; font-size: 0.9em;
        }}
        .bulk-zero {{ color: #cbd5e0; }}
        /* ヒートマップテーブル */
        .heatmap-table {{
            border-collapse: collapse; font-size: 0.82em;
            box-shadow: 0 1px 6px rgba(0,0,0,0.08);
        }}
        .heatmap-table th {{
            background: #4a5568; color: #fff;
            padding: 8px 12px; text-align: center; white-space: nowrap;
        }}
        .heatmap-table .idx-cell {{
            background: #f7fafc; padding: 7px 12px;
            border-bottom: 1px solid #e2e8f0; white-space: nowrap;
            font-size: 0.85em; color: #4a5568;
        }}
        .heatmap-table .wos-cell {{
            text-align: center; padding: 7px 10px; min-width: 54px;
            border-bottom: 1px solid #e2e8f0;
            font-weight: 600; font-size: 0.92em;
        }}
        .heatmap-table .na-cell {{ color: #a0aec0; font-weight: normal; background: #f7fafc; }}
        .heatmap-table tr:hover .wos-cell {{ filter: brightness(0.9); }}
        /* 凡例 */
        .legend {{
            display: flex; align-items: center; gap: 8px;
            font-size: 0.82em; color: #718096; margin-bottom: 14px;
        }}
        .legend-bar {{
            width: 180px; height: 14px; border-radius: 4px;
            background: linear-gradient(to right, #ff8080, #ffffff, #8080ff);
            border: 1px solid #e2e8f0;
        }}
        /* 全社サマリーテーブル */
        .summary-table {{
            border-collapse: collapse; width: 100%;
            box-shadow: 0 1px 6px rgba(0,0,0,0.08); border-radius: 8px; overflow: hidden;
        }}
        .summary-table th {{
            background: #553c9a; color: #fff;
            padding: 10px 14px; text-align: left; font-size: 0.88em; font-weight: 600;
        }}
        .summary-table td {{
            border-bottom: 1px solid #e2e8f0; padding: 8px 14px; font-size: 0.87em;
        }}
        .summary-table tr:hover td {{ background: #faf5ff; }}
        .no-data {{ color: #a0aec0; font-style: italic; }}
        .overflow-x {{ overflow-x: auto; }}
        .st-badge {{
            display: inline-block; background: #fef3c7; color: #92400e;
            border: 1px solid #f59e0b; border-radius: 4px;
            padding: 2px 10px; font-size: 0.82em; margin-left: 8px;
        }}
    </style>
    <script>
        function showTab(tabId, el) {{
            document.querySelectorAll('.tab-content').forEach(e => e.classList.remove('active'));
            document.querySelectorAll('.tab').forEach(e => e.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            el.classList.add('active');
        }}
    </script>
</head>
<body>
    <h1>WOS アイテム集約レポート</h1>
    <p class="meta">生成日時: {generated_at}</p>

    <div class="kpi-bar">
        <div class="kpi"><div class="label">対象SKU数</div><div class="value">{sku_count}</div></div>
        <div class="kpi"><div class="label">対象店舗数</div><div class="value">{store_count}</div></div>
        <div class="kpi priority"><div class="label">⭐ 優先集約件数</div><div class="value">{priority_count}</div></div>
        <div class="kpi"><div class="label">📦 通常集約件数</div><div class="value">{normal_count}</div></div>
    </div>

    <details class="logic-box">
        <summary>計算ロジックについて（クリックで展開）</summary>
        <ul>
            <li><strong>WOS（週数）</strong>: 現在在庫 ÷ 直近4週の平均販売数</li>
            <li><strong>消化率</strong>: 全期間累計売上数 ÷ 発注数（総計列）× 100</li>
            <li><strong>⭐ 優先集約</strong>: 消化率 ≥ {threshold:.0f}% の SKU（完売を狙う）</li>
            <li><strong>📦 通常集約</strong>: 消化率 &lt; {threshold:.0f}% または消化率データなし</li>
            <li><strong>BULK在庫</strong>: 倉庫在庫数（⚠️ 1以上の場合は黄色強調表示：店舗間移動より倉庫出荷を優先検討）</li>
            <li><strong>出荷前WOS → 出荷後WOS</strong>: 移動前の出荷元WOS → 移動後（在庫減）の予測WOS</li>
            <li><strong>受入前WOS → 受入後WOS</strong>: 移動前の受入先WOS → 移動後（在庫増）の予測WOS</li>
            <li><strong>出荷候補</strong>: SKUの全店舗平均WOSより高い店舗（在庫余剰）</li>
            <li><strong>受入候補</strong>: SKUの全店舗平均WOSより低い店舗（在庫不足）</li>
            <li><strong style="color:#c05621;">🟠 オレンジ左枠の行（フォールバックWOS）</strong>: 直近4週売上なしだが全期間に実績あり。全期間平均週販でWOSを補完して移動推奨に含めた行。理由列に「(参考)」と表示。</li>
        </ul>
    </details>

    <div class="tabs">
        <button class="tab priority-tab active" onclick="showTab('tab-priority', this)">⭐ 優先集約リスト<span class="st-badge">{st_badge}</span></button>
        <button class="tab" onclick="showTab('tab-normal', this)">📦 通常集約リスト</button>
        <button class="tab" onclick="showTab('tab-summary', this)">📊 全社サマリー</button>
        <button class="tab" onclick="showTab('tab-wos', this)">🌡 WOS ヒートマップ</button>
    </div>

    <div id="tab-priority" class="tab-content active">
        <h2 style="font-size:1.1em;color:#c05621;margin-bottom:8px;">⭐ 優先集約リスト（消化率 ≥ {threshold:.0f}%）</h2>
        <p style="font-size:0.85em;color:#718096;margin-bottom:12px;">
            完売を狙える商品です。消化率の高い順に並んでいます。
        </p>
        <div class="overflow-x">{priority_html}</div>
    </div>

    <div id="tab-normal" class="tab-content">
        <h2 style="font-size:1.1em;color:#2b6cb0;margin-bottom:8px;">📦 通常集約リスト（消化率 &lt; {threshold:.0f}%）</h2>
        <p style="font-size:0.85em;color:#718096;margin-bottom:12px;">
            在庫の偏りを解消する移動推奨リストです。
        </p>
        <div class="overflow-x">{normal_html}</div>
    </div>

    <div id="tab-summary" class="tab-content">
        <h2 style="font-size:1.1em;color:#553c9a;margin-bottom:8px;">📊 全社サマリー（SKU別統計）</h2>
        <p style="font-size:0.85em;color:#718096;margin-bottom:12px;">
            消化率の高い順に並んでいます。🟢 ≥ 80% / 🟡 60〜79% / 🔴 &lt; 60%
        </p>
        <div class="overflow-x">{summary_html}</div>
    </div>

    <div id="tab-wos" class="tab-content">
        <h2 style="font-size:1.1em;color:#2b6cb0;margin-bottom:8px;">🌡 店舗別 WOS ヒートマップ</h2>
        <div class="legend">
            <span>在庫不足</span>
            <div class="legend-bar"></div>
            <span>在庫余剰</span>
            &nbsp;|&nbsp; セルにカーソルを合わせるとWOS値が表示されます
        </div>
        <div class="overflow-x">{heatmap_html}</div>
    </div>

</body>
</html>
"""
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)

    @staticmethod
    def generate_provisional_matrix_excel(wos_df: pd.DataFrame, move_df: pd.DataFrame, output_path: str = "暫定移動明細.xlsx"):
        """
        作業用マトリクスExcel（商品情報＋在庫＋ORDER＋在庫日数、視覚的強化版）を出力する
        """
        print(f"暫定版マトリクスExcelを生成しています: {output_path}")

        if wos_df.empty:
            print("  [警告] WOSデータが空のため、ファイル生成をスキップしました。")
            return

        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from src.item_allocator import ItemAllocator

        # ヘッダー表示用店舗名マッピング
        def _display_store_name(s: str) -> str:
            st = str(s)
            if 'NODE' in st:
                return 'TOKYO NODE'
            elif 'TOKYO' in st:
                return 'TOKYO'
            elif 'ルクア' in st:
                return 'ルクア大阪'
            elif '名古屋' in st:
                return '名古屋'
            elif '京王新宿' in st:
                return '京王新宿'
            elif '大丸心斎橋' in st:
                return '大丸心斎橋'
            elif '玉川高島屋' in st:
                return '玉川高島屋'
            elif 'HUTTE' in st or 'ヒュッテ' in st or 'HUT' in st:
                return 'APPORO HUT'
            elif 'NARITA' in st:
                return 'NARITA'
            return st

        # 1. 商品マスター情報の抽出（消化率・BULK在庫含む）
        sku_cols = ['sku', 'item_name']
        if 'color_name' in wos_df.columns:
            sku_cols.append('color_name')
        has_st = 'sell_through' in wos_df.columns
        if has_st:
            sku_cols.append('sell_through')
        has_bulk = 'bulk_stock' in wos_df.columns
        if has_bulk:
            sku_cols.append('bulk_stock')

        item_master = wos_df[sku_cols].drop_duplicates('sku').set_index('sku')

        # 消化率降順でソート（消化率の高い順に移動検討できるようにする）
        if has_st:
            item_master = item_master.sort_values(by='sell_through', ascending=False, na_position='last')

        # 2. 店舗リストの決定（受入優先順位でソートし、9店舗に限定）
        all_stores = [s for s in wos_df['store'].dropna().unique() if ItemAllocator._get_store_priority_rank(s) < 99]
        all_stores.sort(key=lambda s: ItemAllocator._get_store_priority_rank(s))
        display_stores = [_display_store_name(s) for s in all_stores]

        # 3. 在庫エリア（各店舗の現在庫）
        stock_pivot = wos_df.pivot_table(
            index='sku',
            columns='store',
            values='stock_qty',
            aggfunc='sum'
        ).fillna(0).astype(int)
        stock_pivot = stock_pivot.reindex(index=item_master.index, columns=all_stores, fill_value=0)

        # 4. ORDERエリア（各店舗の移動推奨数: 出庫=-, 受入=+）
        if not move_df.empty:
            ship_df = move_df[['sku', 'shipper', 'move_qty']].copy()
            ship_df['order_qty'] = -ship_df['move_qty']
            ship_df = ship_df.rename(columns={'shipper': 'store'})[['sku', 'store', 'order_qty']]

            recv_df = move_df[['sku', 'receiver', 'move_qty']].copy()
            recv_df['order_qty'] = recv_df['move_qty']
            recv_df = recv_df.rename(columns={'receiver': 'store'})[['sku', 'store', 'order_qty']]

            order_combined = pd.concat([ship_df, recv_df], ignore_index=True)
            order_pivot = order_combined.pivot_table(
                index='sku',
                columns='store',
                values='order_qty',
                aggfunc='sum'
            ).fillna(0).astype(int)
        else:
            order_pivot = pd.DataFrame(0, index=item_master.index, columns=all_stores)

        order_pivot = order_pivot.reindex(index=item_master.index, columns=all_stores, fill_value=0)

        # 5. 在庫日数（WOS）エリア
        wos_pivot = wos_df.pivot_table(
            index='sku',
            columns='store',
            values='wos',
            aggfunc='mean'
        ).round(1)
        wos_pivot = wos_pivot.reindex(index=item_master.index, columns=all_stores)

        # 在庫があるSKUまたは移動推奨があるSKUを対象にする
        has_stock = (stock_pivot.sum(axis=1) > 0)
        has_move = (order_pivot.abs().sum(axis=1) > 0)
        target_skus = item_master.index[has_stock | has_move]

        if len(target_skus) == 0:
            target_skus = item_master.index

        # ワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'order'
        ws.views.sheetView[0].showGridLines = True

        # スタイル定義
        font_header = Font(name='Meiryo UI', size=9, bold=True)
        font_data = Font(name='Meiryo UI', size=9)
        font_bold = Font(name='Meiryo UI', size=9, bold=True)
        font_gray = Font(name='Meiryo UI', size=9, color='A6A6A6') # ゼロ用グレー文字

        # 消化率用スタイル
        fill_st_high = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') # 緑
        font_st_high = Font(name='Meiryo UI', size=9, bold=True, color='155724')
        fill_st_mid = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 黄
        font_st_mid = Font(name='Meiryo UI', size=9, color='856404')
        fill_st_low = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 赤
        font_st_low = Font(name='Meiryo UI', size=9, color='721C24')

        # BULK在庫用スタイル
        fill_bulk_pos = PatternFill(start_color='FFEAA7', end_color='FFEAA7', fill_type='solid') # 薄オレンジ
        font_bulk_pos = Font(name='Meiryo UI', size=9, bold=True, color='B7791F')

        # ヘッダー背景色
        fill_info_h = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')   # 薄青
        fill_stock_h = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 薄緑
        fill_order_h = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 薄黄
        fill_wos_h = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')    # 薄紫

        fill_stock_cell = PatternFill(start_color='F2F9F0', end_color='F2F9F0', fill_type='solid')
        fill_wos_cell = PatternFill(start_color='FAF5FF', end_color='FAF5FF', fill_type='solid')
        fill_plus = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')     # 入庫（緑）
        fill_minus = PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid')    # 出庫（薄橙）

        # WOSヒートマップ用スタイル
        fill_wos_low = PatternFill(start_color='FCE8E6', end_color='FCE8E6', fill_type='solid')  # 在庫逼迫（薄赤）
        font_wos_low = Font(name='Meiryo UI', size=9, bold=True, color='C5221F')
        fill_wos_high = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid') # 在庫過剰（薄青）
        font_wos_high = Font(name='Meiryo UI', size=9, bold=True, color='1A73E8')

        thin_side = Side(border_style='thin', color='D9D9D9')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # ヘッダー作成
        headers_row1 = ['商品情報', '商品情報']
        headers_row2 = ['商品コード', '商品名']
        if 'color_name' in sku_cols:
            headers_row1.append('商品情報')
            headers_row2.append('カラー')
        if has_st:
            headers_row1.append('商品情報')
            headers_row2.append('消化率(%)')
        if has_bulk:
            headers_row1.append('商品情報')
            headers_row2.append('BULK在庫')

        num_info_cols = len(headers_row1)
        num_stores = len(all_stores)

        # ウィンドウ枠固定（商品情報列の右側で固定）
        freeze_col_letter = get_column_letter(num_info_cols + 1)
        ws.freeze_panes = f'{freeze_col_letter}3'

        headers_row1.extend(['在庫'] * num_stores)
        headers_row2.extend(display_stores)

        headers_row1.extend(['ORDER'] * num_stores)
        headers_row2.extend(display_stores)

        headers_row1.extend(['集約後在庫'] * num_stores)
        headers_row2.extend(display_stores)

        headers_row1.extend(['在庫日数'] * num_stores)
        headers_row2.extend(display_stores)

        ws.append(headers_row1)
        ws.append(headers_row2)

        # ヘッダースタイル適用
        for col_idx in range(1, len(headers_row1) + 1):
            c1 = ws.cell(row=1, column=col_idx)
            c2 = ws.cell(row=2, column=col_idx)
            c1.font = font_header
            c2.font = font_header
            c1.border = border_all
            c2.border = border_all
            c1.alignment = align_center
            c2.alignment = align_center

            fill_post_stock_h = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            if col_idx <= num_info_cols:
                c1.fill = fill_info_h
                c2.fill = fill_info_h
            elif col_idx <= num_info_cols + num_stores:
                c1.fill = fill_stock_h
                c2.fill = fill_stock_h
            elif col_idx <= num_info_cols + num_stores * 2:
                c1.fill = fill_order_h
                c2.fill = fill_order_h
            elif col_idx <= num_info_cols + num_stores * 3:
                c1.fill = fill_post_stock_h
                c2.fill = fill_post_stock_h
            else:
                c1.fill = fill_wos_h
                c2.fill = fill_wos_h

        # データ行の追加
        for sku in target_skus:
            name_val = str(item_master.loc[sku, 'item_name'])
            bulk_val = int(item_master.loc[sku, 'bulk_stock']) if (has_bulk and pd.notna(item_master.loc[sku, 'bulk_stock'])) else 0

            # BULK在庫がある商品は商品名に📦アイコン付与
            if bulk_val > 0:
                name_val = f"📦 {name_val}"

            row_data = [sku, name_val]
            if 'color_name' in sku_cols:
                row_data.append(item_master.loc[sku, 'color_name'])
            if has_st:
                st_val = item_master.loc[sku, 'sell_through']
                row_data.append(round(st_val, 1) if pd.notna(st_val) else '')
            if has_bulk:
                row_data.append(bulk_val)

            row_data.extend(stock_pivot.loc[sku].values)
            row_data.extend(order_pivot.loc[sku].values)

            current_row = ws.max_row + 1
            post_stock_formulas = []
            for i in range(num_stores):
                stock_col = get_column_letter(num_info_cols + i + 1)
                order_col = get_column_letter(num_info_cols + num_stores + i + 1)
                post_stock_formulas.append(f"={stock_col}{current_row}+{order_col}{current_row}")
            row_data.extend(post_stock_formulas)

            # WOS値（NaNは空白）
            wos_vals = [w if pd.notna(w) else '' for w in wos_pivot.loc[sku].values]
            row_data.extend(wos_vals)

            ws.append(row_data)

        # データ行のスタイル適用
        st_col_idx = 4 if has_st else -1
        bulk_col_idx = 5 if (has_st and has_bulk) else (4 if has_bulk else -1)

        for r_idx in range(3, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = font_data
                cell.border = border_all

                # 商品情報エリア (A〜E列)
                if c_idx <= num_info_cols:
                    if c_idx == 2:
                        cell.alignment = align_left   # 商品名
                    elif c_idx == st_col_idx:
                        cell.alignment = align_right  # 消化率(%)
                        val = cell.value
                        if pd.notna(val) and isinstance(val, (int, float)):
                            cell.number_format = '0.0"%"'
                            if val >= 80.0:
                                cell.fill = fill_st_high
                                cell.font = font_st_high
                            elif val >= 60.0:
                                cell.fill = fill_st_mid
                                cell.font = font_st_mid
                            else:
                                cell.fill = fill_st_low
                                cell.font = font_st_low
                    elif c_idx == bulk_col_idx:
                        cell.alignment = align_right  # BULK在庫
                        val = cell.value
                        if pd.notna(val) and isinstance(val, (int, float)):
                            if val > 0:
                                cell.fill = fill_bulk_pos
                                cell.font = font_bulk_pos
                            else:
                                cell.font = font_gray
                    else:
                        cell.alignment = align_center

                # 在庫エリア（ゼロはグレーアウト）
                elif c_idx <= num_info_cols + num_stores:
                    cell.alignment = align_right
                    cell.fill = fill_stock_cell
                    val = cell.value
                    if isinstance(val, (int, float)):
                        if val == 0:
                            cell.font = font_gray
                        else:
                            cell.font = font_bold

                # ORDERエリア
                elif c_idx <= num_info_cols + num_stores * 2:
                    cell.alignment = align_right
                    val = cell.value
                    if isinstance(val, (int, float)):
                        if val > 0:
                            cell.fill = fill_plus
                            cell.font = font_bold
                        elif val < 0:
                            cell.fill = fill_minus
                            cell.font = font_bold
                        else:
                            cell.font = font_gray

                # 集約後在庫エリア
                elif c_idx <= num_info_cols + num_stores * 3:
                    cell.alignment = align_right
                    cell.font = font_bold

                # 在庫日数エリア（WOSヒートマップ）
                else:
                    cell.alignment = align_right
                    cell.fill = fill_wos_cell
                    val = cell.value
                    if pd.notna(val) and isinstance(val, (int, float)):
                        cell.number_format = '0.0'
                        if val < 2.0:
                            cell.fill = fill_wos_low
                            cell.font = font_wos_low
                        elif val > 8.0:
                            cell.fill = fill_wos_high
                            cell.font = font_wos_high

        # 列幅設定
        for c_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c_idx)
            if c_idx == 1:
                ws.column_dimensions[col_letter].width = 18  # 商品コード
            elif c_idx == 2:
                ws.column_dimensions[col_letter].width = 32  # 商品名
            elif c_idx <= num_info_cols:
                ws.column_dimensions[col_letter].width = 14  # カラー / 消化率 / BULK
            else:
                ws.column_dimensions[col_letter].width = 12  # 店舗列

        # 条件付き書式
        from openpyxl.formatting.rule import FormulaRule
        start_letter = get_column_letter(num_info_cols + num_stores + 1)
        end_letter = get_column_letter(num_info_cols + num_stores * 2)

        # 1. ORDER列の合計が0でない場合、A列（商品コード）を赤くハイライト（不一致警告）
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # 薄い赤
        red_font = Font(name='Meiryo UI', size=9, bold=True, color='9C0006')               # 濃い赤文字
        rule_a = FormulaRule(
            formula=[f"SUM(${start_letter}3:${end_letter}3)<>0"],
            stopIfTrue=True,
            fill=red_fill,
            font=red_font
        )
        ws.conditional_formatting.add(f"A3:A{ws.max_row}", rule_a)

        # 2. ORDER列に数字（0以外）が入っている場合、B列（商品名）を黄色にハイライト（移動対象の視覚化）
        yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # 薄い黄
        yellow_font = Font(name='Meiryo UI', size=9, bold=True)
        rule_b = FormulaRule(
            formula=[f'COUNTIF(${start_letter}3:${end_letter}3, "<>0")>0'],
            stopIfTrue=True,
            fill=yellow_fill,
            font=yellow_font
        )
        ws.conditional_formatting.add(f"B3:B{ws.max_row}", rule_b)

        # 3. 集約後在庫の色付け（マイナスは赤、受入は緑、出荷はピンク、0はグレー）
        ps_start = get_column_letter(num_info_cols + num_stores * 2 + 1)
        ps_end = get_column_letter(num_info_cols + num_stores * 3)
        ord_start = get_column_letter(num_info_cols + num_stores + 1)
        ps_range = f"{ps_start}3:{ps_end}{ws.max_row}"

        rule_ps_red = FormulaRule(formula=[f"{ps_start}3<0"], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), font=Font(name='Meiryo UI', size=9, bold=True, color='9C0006'), stopIfTrue=True)
        ws.conditional_formatting.add(ps_range, rule_ps_red)
        
        rule_ps_green = FormulaRule(formula=[f"{ord_start}3>0"], fill=PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'), font=Font(name='Meiryo UI', size=9, bold=True, color='006100'), stopIfTrue=True)
        ws.conditional_formatting.add(ps_range, rule_ps_green)
        
        rule_ps_pink = FormulaRule(formula=[f"{ord_start}3<0"], fill=PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), font=Font(name='Meiryo UI', size=9, bold=True, color='C65911'), stopIfTrue=True)
        ws.conditional_formatting.add(ps_range, rule_ps_pink)
        
        rule_ps_gray = FormulaRule(formula=[f"{ps_start}3=0"], fill=PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'), font=Font(name='Meiryo UI', size=9, color='A6A6A6'), stopIfTrue=True)
        ws.conditional_formatting.add(ps_range, rule_ps_gray)

        try:
            wb.save(output_path)
            print(f"  -> {output_path} 出力完了 ({len(target_skus)} 行)")
        except PermissionError:
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = f"暫定移動明細_{ts}.xlsx"
            print(f"  [警告] '{output_path}' がExcelで開かれているため上書きできませんでした。")
            print(f"         代替ファイル '{fallback_path}' として保存します。")
            wb.save(fallback_path)
            print(f"  -> {fallback_path} 出力完了 ({len(target_skus)} 行)")

