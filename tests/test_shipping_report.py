"""
tests/test_shipping_report.py
-----------------------------
WOS計算、アイテム移動最適化、店舗出荷レポート生成のテストスイート。
"""

import os
import sys
import unittest
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.create_shipping_report import (
    _is_yellow_fill,
    normalize_store,
    calculate_new_moves,
    ALL_STORES
)
from src.wos_calculator import WOSCalculator
from src.item_allocator import ItemAllocator
from src.reporter import Reporter


class TestStoreExclusion(unittest.TestCase):
    def test_excluded_stores(self):
        """対象外店舗（New Way, ZOZO, ユニフォーム, 丸井, バルク, 6142等）が正しく判定されるか"""
        self.assertTrue(WOSCalculator.is_excluded_store('New Way-A'))
        self.assertTrue(WOSCalculator.is_excluded_store('New Way-B'))
        self.assertTrue(WOSCalculator.is_excluded_store('ZOZO'))
        self.assertTrue(WOSCalculator.is_excluded_store('丸井ｳｪﾌﾞﾁｬﾝﾈﾙ'))
        self.assertTrue(WOSCalculator.is_excluded_store('丸井ウェブチャンネル'))
        self.assertTrue(WOSCalculator.is_excluded_store('ﾕﾆﾌｫｰﾑ ﾙｸｱ大阪'))
        self.assertTrue(WOSCalculator.is_excluded_store('ﾕﾆﾌｫｰﾑTOKYO'))
        self.assertTrue(WOSCalculator.is_excluded_store('ﾕﾆﾌｫｰﾑ本社'))
        self.assertTrue(WOSCalculator.is_excluded_store('バルク'))
        self.assertTrue(WOSCalculator.is_excluded_store('BULK'))
        self.assertTrue(WOSCalculator.is_excluded_store('テスト店舗'))
        self.assertTrue(WOSCalculator.is_excluded_store('6142', exclude_stores=['6142']))
        self.assertTrue(WOSCalculator.is_excluded_store('FJALLRAVEN by 3NITY'))

    def test_included_retail_stores(self):
        """正規のリテール実店舗が除外されないか"""
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN by 3NITY TOKYO'))
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN by 3NITY ルクア大阪'))
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN STORE 名古屋ファッションワン'))
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN by 3NITY 大丸心斎橋'))
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN by 3NITY玉川高島屋S・C'))
        self.assertFalse(WOSCalculator.is_excluded_store('FJALLRAVEN POPUP NARITA'))


class TestYellowFillDetection(unittest.TestCase):
    def test_yellow_rgb(self):
        fill_yellow1 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        fill_yellow2 = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        fill_light_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        fill_orange_yellow = PatternFill(start_color='FFEAA7', end_color='FFEAA7', fill_type='solid')
        fill_blue = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
        fill_none = PatternFill(fill_type=None)

        self.assertTrue(_is_yellow_fill(fill_yellow1))
        self.assertTrue(_is_yellow_fill(fill_yellow2))
        self.assertTrue(_is_yellow_fill(fill_light_yellow))
        self.assertTrue(_is_yellow_fill(fill_orange_yellow))
        self.assertFalse(_is_yellow_fill(fill_blue))
        self.assertFalse(_is_yellow_fill(fill_none))
        self.assertFalse(_is_yellow_fill(None))


class TestItemAllocatorEdgeCases(unittest.TestCase):
    def test_wos_avg_nan_handling(self):
        """全ての WOS が NaN の場合にクラッシュせず空 DataFrame を返すか"""
        df = pd.DataFrame([
            {'sku': 'SKU001', 'store': 'TOKYO', 'stock_qty': 0, 'wos': np.nan, 'avg_sales_4w': 0.0},
            {'sku': 'SKU001', 'store': '名古屋', 'stock_qty': 0, 'wos': np.nan, 'avg_sales_4w': 0.0}
        ])
        result = ItemAllocator.allocate(df)
        self.assertTrue(result.empty)

    def test_surplus_deficit_clamping(self):
        """surplus と deficit が float 減算で負数にならず 0.0 にクランプされるか"""
        df = pd.DataFrame([
            {'sku': 'SKU001', 'store': 'TOKYO', 'stock_qty': 10, 'wos': 10.0, 'avg_sales_4w': 1.0},
            {'sku': 'SKU001', 'store': '名古屋', 'stock_qty': 1, 'wos': 1.0, 'avg_sales_4w': 1.0}
        ])
        result = ItemAllocator.allocate(df)
        self.assertFalse(result.empty)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['move_qty'], 5)

    def test_idle_stock_aggregation_single_receiver(self):
        """受入先が1店舗のみの場合、手元に1点残して余剰分を移動するか"""
        df = pd.DataFrame([
            {'sku': 'SKU002', 'store': '大丸心斎橋', 'stock_qty': 2, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan},
            {'sku': 'SKU002', 'store': 'ルクア大阪', 'stock_qty': 0, 'avg_sales_4w': 0.25, 'avg_sales_full': 0.04, 'wos': 0.0}
        ])
        result = ItemAllocator.allocate(df)
        self.assertFalse(result.empty)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['shipper'], '大丸心斎橋')
        self.assertEqual(result.iloc[0]['receiver'], 'ルクア大阪')
        self.assertEqual(result.iloc[0]['move_qty'], 1)

    def test_idle_stock_aggregation_multiple_receivers(self):
        """受入先が2店舗以上の場合、全量移動して各店舗に配分されるか"""
        df = pd.DataFrame([
            {'sku': 'SKU003', 'store': '大丸心斎橋', 'stock_qty': 2, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan},
            {'sku': 'SKU003', 'store': 'ルクア大阪', 'stock_qty': 0, 'avg_sales_4w': 0.25, 'avg_sales_full': 0.04, 'wos': 0.0},
            {'sku': 'SKU003', 'store': '玉川高島屋S・C', 'stock_qty': 0, 'avg_sales_4w': 0.25, 'avg_sales_full': 0.04, 'wos': 0.0}
        ])
        result = ItemAllocator.allocate(df)
        self.assertFalse(result.empty)
        self.assertEqual(len(result), 2)
        total_moved = result['move_qty'].sum()
        self.assertEqual(total_moved, 2)

    def test_idle_stock_single_item_transfer(self):
        """在庫1点の滞留品は受入先1店舗でも全量（1点）移動されるか"""
        df = pd.DataFrame([
            {'sku': 'SKU004', 'store': '大丸心斎橋', 'stock_qty': 1, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan},
            {'sku': 'SKU004', 'store': 'ルクア大阪', 'stock_qty': 0, 'avg_sales_4w': 0.25, 'avg_sales_full': 0.04, 'wos': 0.0}
        ])
        result = ItemAllocator.allocate(df)
        self.assertFalse(result.empty)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['move_qty'], 1)

    def test_no_self_store_move_and_low_sales_receiver_exclusion(self):
        """全期間売上が0.25未満の低実績店が受入先に選出されず、自店舗移動も発生しないか"""
        # 京王新宿: 直近4週=0, 全期間=0.038 (月0.15点程度), 在庫2
        # NARITA: 直近4週=0, 全期間=0, 在庫1
        # SAPPORO: 直近4週=0, 全期間=0, 在庫1
        df = pd.DataFrame([
            {'sku': 'SKU_HAT', 'store': '京王新宿', 'stock_qty': 2, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.038, 'wos': np.nan},
            {'sku': 'SKU_HAT', 'store': 'NARITA', 'stock_qty': 1, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan},
            {'sku': 'SKU_HAT', 'store': 'SAPPORO HUTTE', 'stock_qty': 1, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan}
        ])
        result = ItemAllocator.allocate(df)
        # 有効な受入先（直近売上>0 または 全期間>=0.25）が存在しないため、移動推奨は空になるべき
        self.assertTrue(result.empty, "売上実績のない店舗へ集約移動されたり、自店舗移動が発生してはいけません")

    def test_valid_fallback_receiver_gets_aggregation(self):
        """直近4週=0でも全期間週販が0.25以上の店舗は正しく受入先として選ばれるか"""
        df = pd.DataFrame([
            {'sku': 'SKU_FB', 'store': 'NARITA', 'stock_qty': 1, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.0, 'wos': np.nan},
            {'sku': 'SKU_FB', 'store': 'ルクア大阪', 'stock_qty': 0, 'avg_sales_4w': 0.0, 'avg_sales_full': 0.30, 'wos': np.nan}
        ])
        result = ItemAllocator.allocate(df)
        self.assertFalse(result.empty)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['shipper'], 'NARITA')
        self.assertEqual(result.iloc[0]['receiver'], 'ルクア大阪')
        self.assertEqual(result.iloc[0]['move_qty'], 1)
        self.assertTrue(result.iloc[0]['is_receiver_fallback'])




class TestReporterBulkColumnDynamic(unittest.TestCase):
    def test_apply_excel_formatting_dynamic_bulk(self):
        """BULK在庫の列位置が変わっても安全にハイライトが適用されるか"""
        temp_file = 'test_temp_report.xlsx'
        df = pd.DataFrame([{
            '商品コード': 'SKU001', '商品名': 'Test Item', 'BULK在庫': 5,
            '出荷店舗': 'TOKYO', '受入店舗': '名古屋'
        }])
        raw_df = pd.DataFrame([{
            'sku': 'SKU001', 'item_name': 'Test Item', 'bulk_stock': 5,
            'shipper': 'TOKYO', 'receiver': '名古屋'
        }])

        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='⭐優先集約リスト', index=False)
            Reporter._apply_excel_formatting(writer, '⭐優先集約リスト', raw_df)

        wb = openpyxl.load_workbook(temp_file)
        ws = wb['⭐優先集約リスト']
        cell = ws.cell(row=2, column=3)  # BULK在庫は3列目
        self.assertIsNotNone(cell.fill)
        self.assertTrue(_is_yellow_fill(cell.fill))

        wb.close()
        if os.path.exists(temp_file):
            os.remove(temp_file)


class TestShippingReportGeneratedOutput(unittest.TestCase):
    def test_output_file_exists_and_valid(self):
        """生成された WOS_Report_店舗出荷.xlsx のシート構成と整合性を検証"""
        filepath = 'WOS_Report_店舗出荷.xlsx'
        self.assertTrue(os.path.exists(filepath), f'{filepath} が存在しません')

        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        has_summary = '店舗別出荷サマリー' in wb.sheetnames or '📋店舗別出荷サマリー' in wb.sheetnames
        has_detail = '統合移動明細' in wb.sheetnames or '📋統合集約リスト' in wb.sheetnames
        self.assertTrue(has_summary, '店舗別出荷サマリーシートが見つかりません')
        self.assertTrue(has_detail, '統合移動明細シートが見つかりません')

        for store in ALL_STORES:
            possible_names = [f'出荷_{store}', f'出荷_{store.replace("SAPPORO ", "")}', f'出荷_SAPPORO {store}']
            self.assertTrue(any(sn in wb.sheetnames for sn in possible_names), f'店舗シート（{possible_names}）が見つかりません')

        sum_sheet_title = '店舗別出荷サマリー' if '店舗別出荷サマリー' in wb.sheetnames else '📋店舗別出荷サマリー'
        ws_sum = wb[sum_sheet_title]
        self.assertTrue('店舗間移動 出荷・受入サマリー' in str(ws_sum['B2'].value))
        wb.close()


if __name__ == '__main__':
    unittest.main()
