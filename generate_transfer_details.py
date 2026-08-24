import openpyxl
import pandas as pd
import argparse
import os
import sys
import io

# Windowsのコンソール出力文字化け対策
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def get_store_area(store_name: str) -> str:
    """店舗名から所属エリアを判定する"""
    s = str(store_name).strip()
    if 'NODE' in s:
        return 'KANTO'
    elif 'TOKYO' in s or '東京' in s:
        return 'KANTO'
    elif '新宿' in s:
        return 'KANTO'
    elif '高島屋' in s or '玉川' in s:
        return 'KANTO'
    elif 'NARITA' in s or '成田' in s:
        return 'KANTO'
    elif 'ルクア' in s:
        return 'KANSAI'
    elif '心斎橋' in s or '大丸' in s:
        return 'KANSAI'
    elif '名古屋' in s:
        return 'TOKAI'
    elif 'HUT' in s or 'ヒュッテ' in s:
        return 'HOKKAIDO'
    return 'OTHER'

def get_transfer_cost(shipper: str, receiver: str) -> int:
    """
    2店舗間の配送コスト（距離レベル）を返す（小さいほど低コスト・優先）
    - 1: 同一エリア内（最優先・最安）
    - 2: 隣接エリア間（東海 ⇄ 関東/関西）
    - 3: 遠隔エリア間（関東 ⇄ 関西）
    - 4: 北海道 ⇄ 本州（最長距離）
    """
    a1 = get_store_area(shipper)
    a2 = get_store_area(receiver)

    if a1 == a2 and a1 != 'OTHER':
        return 1

    pair = {a1, a2}
    if pair == {'KANTO', 'TOKAI'} or pair == {'KANSAI', 'TOKAI'}:
        return 2
    elif pair == {'KANTO', 'KANSAI'}:
        return 3
    elif 'HOKKAIDO' in pair:
        return 4
    else:
        return 5

def parse_target_store_from_comment(comment_text: str, candidate_stores: list) -> str:
    """セルコメントから指定された店舗名を判定する"""
    if not comment_text:
        return None
    c = comment_text.strip().lower()

    # 直接マッチ
    for store in candidate_stores:
        if store.lower() in c:
            return store

    # エリア・店舗キーワードマッチ
    keywords = [
        ('node', 'TOKYO NODE'),
        ('tokyo', 'TOKYO'),
        ('東京', 'TOKYO'),
        ('ルクア', 'ルクア大阪'),
        ('心斎橋', '大丸心斎橋'),
        ('大丸', '大丸心斎橋'),
        ('名古屋', '名古屋'),
        ('新宿', '京王新宿'),
        ('京王', '京王新宿'),
        ('高島屋', '玉川高島屋'),
        ('玉川', '玉川高島屋'),
        ('hut', 'APPORO HUT'),
        ('ヒュッテ', 'APPORO HUT'),
        ('narita', 'NARITA'),
        ('成田', 'NARITA'),
    ]
    for kw, target_name in keywords:
        if kw in c:
            for store in candidate_stores:
                if target_name in store or store in target_name:
                    return store
    return None

def main():
    parser = argparse.ArgumentParser(description="作業用マトリクスExcelから配送コスト最小化＆コメント指定を考慮して縦持ちCSVを作成します")
    parser.add_argument("input", help="入力する暫定移動明細Excelのパス")
    parser.add_argument("--output", default="transfer_upload.csv", help="出力するCSVのパス")
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output

    if not os.path.exists(input_path):
        print(f"[エラー] 入力ファイルが見つかりません: {input_path}")
        return

    print(f"[{input_path}] を読み込んでいます（配送コスト最適化モード）...")

    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheet_name = 'order' if 'order' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        # ヘッダー解析（行1: カテゴリ, 行2: 項目名/店舗名）
        row1 = [cell.value for cell in ws[1]]
        row2 = [cell.value for cell in ws[2]]

        sku_col_idx = None
        name_col_idx = None
        color_col_idx = None
        st_col_idx = None
        order_cols = []  # [(col_idx, store_name)]

        for c_idx in range(1, len(row1) + 1):
            cat = str(row1[c_idx - 1] or '').strip()
            item = str(row2[c_idx - 1] or '').strip()

            if item == '商品コード':
                sku_col_idx = c_idx
            elif item == '商品名':
                name_col_idx = c_idx
            elif item == 'カラー':
                color_col_idx = c_idx
            elif '消化率' in item:
                st_col_idx = c_idx
            elif cat.upper() == 'ORDER':
                order_cols.append((c_idx, item))

        if not sku_col_idx:
            print("[エラー] 「商品コード」列が見つかりません。")
            return
        if not order_cols:
            print("[エラー] 「ORDER」カテゴリの列が見つかりません。")
            return

        candidate_stores = [s_name for _, s_name in order_cols]
        print(f"対象店舗数: {len(candidate_stores)} ({', '.join(candidate_stores)})")

        records = []
        unbalanced_rows = 0
        comment_matched_count = 0
        area_matched_count = 0

        # データ行の処理（行3以降）
        for r_idx in range(3, ws.max_row + 1):
            sku_val = ws.cell(row=r_idx, column=sku_col_idx).value
            if not sku_val or str(sku_val).strip() == '':
                continue

            sku = str(sku_val).strip()
            item_name = str(ws.cell(row=r_idx, column=name_col_idx).value or '').strip() if name_col_idx else ''
            color_name = str(ws.cell(row=r_idx, column=color_col_idx).value or '').strip() if color_col_idx else ''

            shippers = []   # [{'store': s, 'qty': int, 'target_store': str or None}]
            receivers = []  # [{'store': r, 'qty': int, 'source_store': str or None}]

            for c_idx, store_name in order_cols:
                cell = ws.cell(row=r_idx, column=c_idx)
                val = cell.value
                comment_text = cell.comment.text if cell.comment else None
                specified_store = parse_target_store_from_comment(comment_text, candidate_stores)

                if val is not None and isinstance(val, (int, float)):
                    val = int(val)
                    if val < 0:
                        shippers.append({
                            'store': store_name,
                            'qty': abs(val),
                            'target_store': specified_store
                        })
                    elif val > 0:
                        receivers.append({
                            'store': store_name,
                            'qty': val,
                            'source_store': specified_store
                        })

            if not shippers and not receivers:
                continue

            total_ship = sum(s['qty'] for s in shippers)
            total_recv = sum(r['qty'] for r in receivers)

            if total_ship != total_recv:
                unbalanced_rows += 1
                if unbalanced_rows <= 5:
                    print(f"  [注意] 出入庫不一致: 行={r_idx}, SKU={sku}, 出庫合計={total_ship}, 入庫合計={total_recv}")

            # ==========================================
            # ステップ1: セルコメントによる手動指定マッチング（最優先）
            # ==========================================
            # 1-1. 出庫側からの指定
            for s in shippers:
                if s['target_store'] and s['qty'] > 0:
                    for r in receivers:
                        if r['store'] == s['target_store'] and r['qty'] > 0:
                            alloc = min(s['qty'], r['qty'])
                            if alloc > 0:
                                records.append({
                                    '移動元店舗': s['store'],
                                    '移動先店舗': r['store'],
                                    '商品コード': sku,
                                    '商品名': item_name,
                                    'カラー': color_name,
                                    '数量': alloc,
                                    'マッチング種別': '📌コメント指定'
                                })
                                s['qty'] -= alloc
                                r['qty'] -= alloc
                                comment_matched_count += alloc

            # 1-2. 受入側からの指定
            for r in receivers:
                if r['source_store'] and r['qty'] > 0:
                    for s in shippers:
                        if s['store'] == r['source_store'] and s['qty'] > 0:
                            alloc = min(s['qty'], r['qty'])
                            if alloc > 0:
                                records.append({
                                    '移動元店舗': s['store'],
                                    '移動先店舗': r['store'],
                                    '商品コード': sku,
                                    '商品名': item_name,
                                    'カラー': color_name,
                                    '数量': alloc,
                                    'マッチング種別': '📌コメント指定'
                                })
                                s['qty'] -= alloc
                                r['qty'] -= alloc
                                comment_matched_count += alloc

            # ==========================================
            # ステップ2: 配送コスト最小化（エリア優先）自動マッチング
            # ==========================================
            # 残りの数量がある出庫・受入ペアの全組み合わせを作成し、コスト昇順でソート
            pair_candidates = []
            for s in shippers:
                if s['qty'] <= 0: continue
                for r in receivers:
                    if r['qty'] <= 0: continue
                    cost = get_transfer_cost(s['store'], r['store'])
                    pair_candidates.append((cost, s, r))

            # コスト昇順（同一エリア=1 ➔ 隣接エリア=2 ➔ 遠隔=3 ➔ 北海道=4）
            pair_candidates.sort(key=lambda x: x[0])

            for cost, s, r in pair_candidates:
                if s['qty'] <= 0 or r['qty'] <= 0:
                    continue
                alloc = min(s['qty'], r['qty'])
                if alloc > 0:
                    area_label = '🚚同一エリア' if cost == 1 else ('🚛隣接エリア' if cost == 2 else '✈️長距離移動')
                    records.append({
                        '移動元店舗': s['store'],
                        '移動先店舗': r['store'],
                        '商品コード': sku,
                        '商品名': item_name,
                        'カラー': color_name,
                        '数量': alloc,
                        'マッチング種別': area_label
                    })
                    s['qty'] -= alloc
                    r['qty'] -= alloc
                    area_matched_count += alloc

        if unbalanced_rows > 0:
            print(f"  [警告] 合計 {unbalanced_rows} 件のSKUで出入庫数が不一致でした（マッチ可能な分のみ変換しました）。")

        if not records:
            print("[情報] 移動対象となる数量（ORDER）が見つかりませんでした。")
            return

        out_df = pd.DataFrame(records)

        # 出力用カラム
        cols = ['移動元店舗', '移動先店舗', '商品コード', '商品名', 'カラー', '数量']
        cols = [c for c in cols if c in out_df.columns]
        out_df_export = out_df[cols].sort_values(by=['移動元店舗', '移動先店舗', '商品コード'])

        out_df_export.to_csv(output_path, index=False, encoding='utf-8-sig')

        # マッチング統計の表示
        print(f"\n================ 変換完了 ================")
        print(f"出力ファイル: {output_path}")
        print(f"総移動点数: {out_df['数量'].sum()} 点 ({len(out_df_export)} 件の明細)")
        if 'マッチング種別' in out_df.columns:
            stats = out_df.groupby('マッチング種別')['数量'].sum()
            for label, count in stats.items():
                print(f"  - {label}: {count} 点")
        print("==========================================")

    except Exception as e:
        print(f"[エラー] 変換中にエラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
